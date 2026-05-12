# -*- coding: utf-8 -*-
import os
import subprocess
import requests
from moviepy.editor import VideoFileClip
from mutagen import File as AudioFile
from fpdf import FPDF
from faster_whisper import WhisperModel
from dotenv import load_dotenv
import json
import base64
import asyncio
from shazamio import Shazam
from PIL import Image, ImageOps
import face_recognition
import numpy as np
import cv2
import shutil
import PyPDF2
from pyrogram import Client
import asyncio
from vision_ocr import recognize_image_file
from document_text import clean_to_single_block, extract_docx_text

# ====== Error helpers: код -> расшифровка и запись в файл ======
import traceback
import json


# --- OpenAI настройки ---
import os

# Основная модель (если не указана в .env — берём gpt-4o-mini)
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")

# Температура генерации (по умолчанию 0.2)
OPENAI_TEMPERATURE = float(os.getenv("OPENAI_TEMPERATURE", "0.2"))

# Максимум токенов на вывод (по умолчанию 1200)
OPENAI_MAX_TOKENS = int(os.getenv("OPENAI_MAX_TOKENS", "1200"))

# Effort для reasoning-моделей (если не задан в .env — используем "medium")
OPENAI_REASONING_EFFORT = os.getenv("OPENAI_REASONING_EFFORT", "medium")

# Список reasoning-моделей (если используете o3-mini, o4-mini и т.п.)
REASONING_MODELS = {"o3-mini", "o4-mini"}

# Прокси (если в .env есть OPENAI_PROXY_URL — то используем)
# --- OpenAI: единая настройка прокси из окружения ---
def _resolve_openai_proxy_url() -> str:
    """
    Берём первый доступный прокси из переменных:
    OPENAI_PROXY_URL / OPENAI_PROXY / HTTPS_PROXY / https_proxy / ALL_PROXY / all_proxy / HTTP_PROXY / http_proxy
    """
    for k in ("OPENAI_PROXY_URL", "OPENAI_PROXY", "HTTPS_PROXY", "https_proxy",
              "ALL_PROXY", "all_proxy", "HTTP_PROXY", "http_proxy"):
        v = os.getenv(k)
        if v:
            return v
    return ""

_OAI_PROXY_URL = _resolve_openai_proxy_url()
PROXIES = {"http": _OAI_PROXY_URL, "https": _OAI_PROXY_URL} if _OAI_PROXY_URL else None





def _write_error_file(path, provider, code, message, explanation, extra=None):
    """
    Записывает человекочитаемый отчёт об ошибке в целевой файл,
    чтобы Flask показал это как «результат».
    """
    lines = []
    lines.append(f"❌ Ошибка {provider}")
    if code:
        lines.append(f"Код: {code}")
    if message:
        lines.append(f"Сообщение: {message}")
    if explanation:
        lines.append(f"Расшифровка: {explanation}")
    if extra:
        lines.append("")
        lines.append("Технические детали:")
        if isinstance(extra, (dict, list)):
            try:
                lines.append(json.dumps(extra, ensure_ascii=False, indent=2))
            except Exception:
                lines.append(str(extra))
        else:
            lines.append(str(extra))
    text = "\n".join(lines).strip() + "\n"
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(text)
    except Exception:
        # Если даже файл не записался — ничего не ломаем
        pass

# Частые коды ошибок Yandex (Cloud/gRPC-style) и краткая расшифровка
_YA_EXPLAIN = {
    "INVALID_ARGUMENT": "Некорректный запрос или параметры (проверьте payload и типы).",
    "FAILED_PRECONDITION": "Нарушено требование к предварительным условиям (конфигурация, состояние ресурса).",
    "OUT_OF_RANGE": "Размер или параметр запроса вне допустимого диапазона.",
    "UNAUTHENTICATED": "Не удалось аутентифицироваться (проверьте IAM-токен/ключ/папку).",
    "PERMISSION_DENIED": "Недостаточно прав для операции (проверьте роли и права папки/сервиса).",
    "NOT_FOUND": "Ресурс не найден (папка, модель или путь).",
    "ABORTED": "Операция прервана (повторите запрос).",
    "RESOURCE_EXHAUSTED": "Превышены лимиты/квоты (скорость, объём).",
    "CANCELLED": "Операция отменена клиентом/сервером.",
    "DATA_LOSS": "Серьёзная ошибка целостности данных на стороне сервиса.",
    "UNKNOWN": "Неизвестная внутренняя ошибка.",
    "INTERNAL": "Внутренняя ошибка сервера.",
    "UNAVAILABLE": "Сервис временно недоступен (повторите запрос позже).",
    "DEADLINE_EXCEEDED": "Превышен таймаут выполнения.",
}

def explain_yandex_error(code: str) -> str:
    if not code:
        return "Произошла ошибка YandexGPT (код не распознан)."
    return _YA_EXPLAIN.get(code, "Неизвестный код. Проверьте запрос и права доступа.")

# Частые коды/типы OpenAI
_OAI_EXPLAIN = {
    "invalid_api_key": "Неверный или неактивный API-ключ.",
    "insufficient_quota": "Недостаточно квоты/средств в аккаунте.",
    "rate_limit_exceeded": "Превышен лимит запросов (подождите и повторите).",
    "context_length_exceeded": "Слишком большой запрос (обрежьте входные данные).",
    "server_error": "Внутренняя ошибка сервера OpenAI.",
    "timeout": "Превышен таймаут запроса.",
    "unsupported_model": "Модель недоступна или отключена.",
    "invalid_request_error": "Неверно сформирован запрос (проверьте поля).",
    "authentication_error": "Ошибка аутентификации (ключ/заголовки).",
    "permission_error": "Недостаточно прав на использование ресурса.",
    "not_found_error": "Ресурс или модель не найдены.",
    "bad_gateway": "Сбой шлюза (повторите запрос).",
    "service_unavailable": "Сервис временно недоступен.",
}

def explain_openai_error(code: str) -> str:
    if not code:
        return "Произошла ошибка OpenAI (код не распознан)."
    return _OAI_EXPLAIN.get(code, "Неизвестный код. Проверьте ключ, модель и формат запроса.")

def _extract_code_and_message_from_response_like(obj):
    """
    Пытаемся вытащить (code, message) из популярных форматов ошибок SDK:
    - {'error': {'code': '...', 'message': '...'}}
    - {'code': '...', 'message': '...'}
    - Exception с полями .code / .message
    - HTTP/gRPC с атрибутами
    """
    code = None
    message = None
    if obj is None:
        return code, message

    # dict / JSON-подобное
    if isinstance(obj, dict):
        if "error" in obj and isinstance(obj["error"], dict):
            code = obj["error"].get("code") or obj["error"].get("type")
            message = obj["error"].get("message")
            return code, message
        code = obj.get("code") or obj.get("status") or obj.get("type")
        message = obj.get("message") or obj.get("detail") or obj.get("error_message")
        return code, message

    # Exception-подобное
    code = getattr(obj, "code", None) or getattr(obj, "status", None) or getattr(obj, "type", None)
    message = getattr(obj, "message", None)
    if message is None:
        # как правило, у Exception есть str(...)
        try:
            message = str(obj)
        except Exception:
            message = None
    return code, message
import re

_PLACEHOLDER_RE = re.compile(
    r"Не удалось получить ответ от OpenAI\s*\(ChatGPT\).*OpenAI_Materials_Log\.json",
    re.IGNORECASE | re.DOTALL
)

def _read_text(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return f.read()
    except Exception:
        return ""

def _write_text(path, text):
    try:
        with open(path, "w", encoding="utf-8") as f:
            f.write(text if text.endswith("\n") else text + "\n")
    except Exception:
        pass


# === FIX: добавьте этот хелпер в run_analysis.py (рядом с _write_text) ===
def _write_json(path: str, data) -> None:
    import os, json
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception as e:
        # не роняем пайплайн, просто логируем в stdout (видно во Flask)
        try:
            print(f"⚠️ _write_json: не удалось записать {path}: {e}")
        except Exception:
            pass



def _extract_human_error_from_openai_log(log_path):
    """
    Пытаемся достать «живую» строку ошибки из лога OpenAI.
    Поддерживаем JSON и «простой текст».
    """
    if not os.path.exists(log_path):
        return None

    # Сначала пробуем разобрать как JSON
    try:
        with open(log_path, "r", encoding="utf-8") as f:
            raw = json.load(f)
        # 1) {"error": {"message": "..."}}
        if isinstance(raw, dict) and isinstance(raw.get("error"), dict):
            for k in ("message", "exception_string", "detail", "error_text", "http_error"):
                if raw["error"].get(k):
                    return str(raw["error"][k]).strip()
        # 2) {"exception": "..."} на верхнем уровне
        if isinstance(raw, dict):
            for k in ("exception", "http_error", "error_text", "detail", "message"):
                if raw.get(k):
                    return str(raw[k]).strip()
        # 3) Список событий
        if isinstance(raw, list):
            for item in raw:
                if isinstance(item, dict):
                    if isinstance(item.get("error"), dict) and item["error"].get("message"):
                        return str(item["error"]["message"]).strip()
                    for k in ("exception", "http_error", "error_text", "detail", "message"):
                        if item.get(k):
                            return str(item[k]).strip()
    except Exception:
        # Если не JSON — читаем как текст
        pass

    # Плоский текст: найдём информативную строку
    text = _read_text(log_path)
    if not text:
        return None

    # Ищем строки с типовыми маркерами
    for line in text.splitlines():
        s = line.strip()
        if not s:
            continue
        if (
            "HTTPError" in s or "Client Error" in s or
            "Too Many Requests" in s or "RateLimit" in s or
            "401" in s or "403" in s or "429" in s or
            "invalid_api_key" in s or "insufficient_quota" in s or
            "context_length_exceeded" in s or "Service Unavailable" in s
        ):
            return s
    # Если ничего — вернём первые 200 символов как fallback
    return text.strip()[:200] or None

def _maybe_replace_openai_placeholder_with_log(out_path, result_dir):
    """
    Если в out_path записана «заглушка», пытаемся заменить её на строку из лога.
    Поддерживаем несколько имён логов.
    """
    cur = _read_text(out_path)
    if not cur or not _PLACEHOLDER_RE.search(cur):
        return  # нет заглушки — ничего не делаем

    # Возможные имена логов
    candidates = [
        os.path.join(result_dir, "OpenAI_Materials_Log.json"),
        os.path.join(result_dir, "OpenAI_Materials_Log.txt"),
        os.path.join(result_dir, "OpenAI_Materials_Log.log"),
    ]
    for lp in candidates:
        msg = _extract_human_error_from_openai_log(lp)
        if msg:
            _write_text(out_path, f"Ошибка OpenAI: {msg}")
            return
    # Лога нет/пуст — оставим как есть









# Загрузка переменных окружения
# безопасное чтение окружения (ничего не падает, если Telegram не настроен)
load_dotenv()
preview_dir = os.environ["PREVIEW_PATH"]
yandex_api_key = os.environ["YANDEX_API_KEY"]
yandex_folder_id = os.environ["YANDEX_FOLDER_ID"]

def _get_int_env(name, default=None):
    v = os.getenv(name, "")
    try:
        return int(v) if str(v).strip() else default
    except Exception:
        return default

telegram_api_id = _get_int_env("TELEGRAM_API_ID")
telegram_api_hash = os.getenv("TELEGRAM_API_HASH") or ""
TELEGRAM_ENABLED = bool(telegram_api_id and telegram_api_hash)


# ===== НОВОЕ: Онлайн-лог и статус для Flask (SSE) =====
import sys, time, json

def _init_live_log_and_status(preview_dir_path: str):
    """Создаёт Clipchecker_materials, включает tee для stdout/stderr в clipchecker.log и пишет стартовый статус."""
    res_dir = os.path.join(preview_dir_path, "Clipchecker_materials")
    os.makedirs(res_dir, exist_ok=True)
    log_path = os.path.join(res_dir, "clipchecker.log")
    status_path = os.path.join(res_dir, "run_status.json")

    # Линейно-буферизованный файл: запись строки сразу видна в браузере
    _log_file = open(log_path, "a", encoding="utf-8", buffering=1)

    class _Tee:
        def __init__(self, *streams):
            self.streams = streams
        def write(self, data):
            for s in self.streams:
                try:
                    s.write(data)
                    s.flush()
                except Exception:
                    pass
        def flush(self):
            for s in self.streams:
                try:
                    s.flush()
                except Exception:
                    pass

    # Перехватываем stdout/stderr на всё время ранa
    sys.stdout = _Tee(sys.stdout, _log_file)
    sys.stderr = _Tee(sys.stderr, _log_file)

    # Стартовый баннер и статус
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n================ Clipchecker: START {ts} ================\n", flush=True)
    try:
        with open(status_path, "w", encoding="utf-8") as f:
            json.dump({"state": "running", "stage": "start", "ts": ts}, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    return res_dir, log_path, status_path

# Инициализация
result_dir, _LOG_PATH, _STATUS_PATH = _init_live_log_and_status(preview_dir)
# ===== /НОВОЕ =====






# Получение пути к папке документов
documents_dir = os.environ.get("DOCUMENTS_PATH")
if not documents_dir or not os.path.exists(documents_dir):
    raise FileNotFoundError("❌ DOCUMENTS_PATH не задан или путь не существует")

# Поиск видеофайла
video_files = [f for f in os.listdir(preview_dir) if f.lower().endswith((".mxf", ".mp4"))]
if not video_files:
    raise FileNotFoundError(f"В папке {preview_dir} не найдено ни одного .mp4 файла")
elif len(video_files) > 1:
    print(f"⚠️ Найдено несколько видеофайлов. Используется: {video_files[0]}")
video_file = video_files[0]
video_path = os.path.join(preview_dir, video_file)

# Подготовка директорий
result_dir = os.path.join(preview_dir, "Clipchecker_materials")
frames_dir = os.path.join(result_dir, "frames")
pdf_frames_dir = os.path.join(result_dir, "frames_pdf_original")
os.makedirs(frames_dir, exist_ok=True)
os.makedirs(pdf_frames_dir, exist_ok=True)
print(f"📂 Временная папка: {frames_dir}")



# Создание списка файлов из DOCUMENTS_PATH
documents_list_path = os.path.join(result_dir, "Document_Files.txt")
with open(documents_list_path, "w", encoding="utf-8") as f:
    for root, _, files in os.walk(documents_dir):
        for filename in files:
            name, ext = os.path.splitext(filename)
            ext = ext.lstrip(".")
            if name and ext:
                f.write(f"{name}.{ext}\n")
            elif name:
                f.write(f"{name}\n")
print(f"📄 Список файлов из DOCUMENTS_PATH сохранён: {documents_list_path}")

# Извлечение кадров
print("📥 Извлекаем кадры через ffmpeg...")
subprocess.run([
    "ffmpeg", "-i", video_path, "-vf", "fps=1",
    os.path.join(pdf_frames_dir, "frame_%03d.jpg"),
    "-hide_banner", "-loglevel", "error"
])
print("✅ Кадры без рамок сохранены в папке для PDF")

# Генерация PDF из оригинальных кадров (без рамок)
pdf_path = os.path.join(result_dir, "frames.pdf")
pdf = FPDF()
for file in sorted(os.listdir(pdf_frames_dir)):
    if file.endswith(".jpg"):
        img_path = os.path.join(pdf_frames_dir, file)
        pdf.add_page()
        pdf.image(img_path, x=10, y=10, w=190)
pdf.output(pdf_path)
print(f"📄 PDF создан: {pdf_path}")

# Копирование кадров в папку для добавления рамок
for file in sorted(os.listdir(pdf_frames_dir)):
    if file.endswith(".jpg"):
        shutil.copy(os.path.join(pdf_frames_dir, file), os.path.join(frames_dir, file))

# Добавление зелёной рамки к кадрам
from PIL import ImageDraw

def add_green_frame(image_path):
    img = Image.open(image_path)
    w, h = img.size
    border_color = (0, 255, 0)
    draw = ImageDraw.Draw(img)
    # Отступы: 10% сверху/снизу, 5% слева/справа
    top = int(h * 0.05)
    bottom = h - int(h * 0.05)
    left = int(w * 0.10)
    right = w - int(w * 0.10)
    border_width = 3
    for i in range(border_width):
        draw.rectangle([left - i, top - i, right + i, bottom + i], outline=border_color)
    img.save(image_path)

print("🎨 Добавляем зелёные рамки с отступами...")
for file in sorted(os.listdir(frames_dir)):
    if file.endswith(".jpg"):
        add_green_frame(os.path.join(frames_dir, file))
print("✅ Рамки добавлены")

# ----- Остальная часть программы -----

# Озвучка через FasterWhisper
def extract_subtitles(video_path, output_path):
    model = WhisperModel("base", device="cpu", compute_type="int8")
    segments, _ = model.transcribe(video_path, language="ru")
    text = "\n".join([seg.text for seg in segments])
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(text or "Speech not recognized")

speech_txt = os.path.join(result_dir, "Speech.txt")
extract_subtitles(video_path, speech_txt)
print(f"🗣 Озвучка сохранена: {speech_txt}")

# Музыка через mutagen
def extract_music_info(video_path, output_path):
    try:
        clip = VideoFileClip(video_path)
        audio_path = os.path.join(result_dir, "temp_audio.mp3")
        clip.audio.write_audiofile(audio_path, verbose=False, logger=None)
        audio = AudioFile(audio_path)
        artist = audio.get("TPE1")
        title = audio.get("TIT2")
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("Музыка есть\n")
            if artist and title:
                f.write(f"{artist.text[0]} - {title.text[0]}")
            else:
                f.write("Не удалось распознать исполнителя")
        os.remove(audio_path)
    except:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write("Музыка не найдена\nНе удалось извлечь данные")

music_txt = os.path.join(result_dir, "Music.txt")
extract_music_info(video_path, music_txt)
print(f"🎶 Информация о музыке сохранена: {music_txt}")

# OCR через Yandex Vision
ocr_txt = os.path.join(result_dir, "Frame_OCR.txt")
combined_txt = os.path.join(result_dir, "All_Frames_Text.txt")
ocr_log = os.path.join(result_dir, "OCR_Log.json")
with open(ocr_txt, "w", encoding="utf-8") as output, open(combined_txt, "w", encoding="utf-8") as combined, open(ocr_log, "w", encoding="utf-8") as log:
    all_results = {}
    for idx, file in enumerate(sorted(os.listdir(frames_dir))):
        if file.endswith(".jpg"):
            path = os.path.join(frames_dir, file)
            result = recognize_image_file(
                path,
                api_key=yandex_api_key,
                folder_id=yandex_folder_id,
                attempts=int(os.getenv("YANDEX_VISION_RETRIES", "3")),
                timeout=int(os.getenv("YANDEX_VISION_TIMEOUT", "90")),
            )
            all_results[file] = result.raw
            if result.ok:
                frame_text = result.text
                output.write(f"{file}: {frame_text or '[пусто]'}\n")
                combined.write(f"Кадр {idx+1}:\n{frame_text or '[пусто]'}\n\n")
            else:
                message = result.error or "unknown OCR error"
                output.write(f"[Ошибка обработки кадра {file}]: {message}\n")
                combined.write(f"Кадр {idx+1}:\n[Ошибка чтения текста: {message}]\n\n")
                all_results[file] = {
                    "error": message,
                    "attempts": result.attempts,
                    "raw": result.raw,
                }
                print(f"⚠️ OCR кадра {file} не выполнен после {result.attempts} попыток: {message}")
    json.dump(all_results, log, ensure_ascii=False, indent=2)
print("🔤 Текст с кадров сохранён")

# Озвучка через Yandex SpeechKit с логом
def transcribe_yandex_speechkit(video_path, output_path):
    try:
        audio_ogg = os.path.join(result_dir, "speech_ogg.ogg")
        subprocess.run([
            "ffmpeg", "-y", "-i", video_path, "-vn", "-ar", "48000", "-ac", "1",
            "-c:a", "libopus", audio_ogg,
            "-hide_banner", "-loglevel", "error"
        ])
        if not os.path.exists(audio_ogg) or os.path.getsize(audio_ogg) < 5000:
            raise RuntimeError("Файл с озвучкой не создан или слишком мал")
        with open(audio_ogg, "rb") as f:
            data = f.read()
        headers = {
            "Authorization": f"Api-Key {yandex_api_key}",
            "Content-Type": "application/octet-stream",
        }
        params = {
            "lang": "ru-RU",
            "folderId": yandex_folder_id,
            "format": "oggopus",
            "sampleRateHertz": 48000
        }
        url = "https://stt.api.cloud.yandex.net/speech/v1/stt:recognize"
        response = requests.post(url, headers=headers, params=params, data=data)
        result = response.json()
        with open(os.path.join(result_dir, "speechkit_raw_response.json"), "w", encoding="utf-8") as debug_file:
            json.dump(result, debug_file, ensure_ascii=False, indent=2)
        if "result" in result:
            text = result["result"]
        elif "error_code" in result:
            text = f"[Ошибка SpeechKit] {result['error_code']}: {result.get('message', '')}"
        else:
            text = f"[Неизвестный ответ от API]:\n{json.dumps(result, ensure_ascii=False, indent=2)}"
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(text)
        print(f"🗣 Озвучка Yandex SpeechKit сохранена: {output_path}")
        os.remove(audio_ogg)
    except Exception as e:
        print(f"❌ Ошибка распознавания речи Yandex: {e}")

speechkit_txt = os.path.join(result_dir, "Speech_Yandex.txt")
transcribe_yandex_speechkit(video_path, speechkit_txt)

# Музыка через Shazamio
async def recognize_with_shazam(video_path, output_path):
    try:
        audio_snippet = os.path.join(result_dir, "shazam_sample.mp3")
        subprocess.run([
            "ffmpeg", "-i", video_path, "-t", "15", "-q:a", "0", "-map", "a", audio_snippet,
            "-hide_banner", "-loglevel", "error"
        ])
        if not os.path.exists(audio_snippet) or os.path.getsize(audio_snippet) < 10000:
            raise RuntimeError("Аудиофайл для Shazam слишком короткий или пуст")
        shazam = Shazam()
        out = await shazam.recognize_song(audio_snippet)
        track = out.get("track", {})
        title = track.get("title", "Неизвестно")
        subtitle = track.get("subtitle", "Неизвестно")
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(f"{subtitle} - {title}")
        print(f"🎧 Результат Shazam сохранён: {output_path}")
    except Exception as e:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(f"Не удалось распознать музыку через Shazam\n{str(e)}")
        print(f"❌ Ошибка Shazam: {e}")

music_shazam_txt = os.path.join(result_dir, "Music_Shazam.txt")
asyncio.run(recognize_with_shazam(video_path, music_shazam_txt))

# Извлечение лиц (оставлено без изменений)
def is_frontal_face(landmarks):
    try:
        left_eye = landmarks["left_eye"]
        right_eye = landmarks["right_eye"]
        nose_tip = landmarks["nose_tip"][2]
        eye_center_x = (np.mean([p[0] for p in left_eye]) + np.mean([p[0] for p in right_eye])) / 2
        symmetry = abs(nose_tip[0] - eye_center_x)
        return symmetry < 20
    except:
        return True

def estimate_sharpness(image):
    gray = cv2.cvtColor(image, cv2.COLOR_RGB2GRAY)
    return cv2.Laplacian(gray, cv2.CV_64F).var()

def extract_faces_from_frames(frames_dir, result_dir, tolerance=0.6):
    faces_dir = os.path.join(result_dir, "faces")
    os.makedirs(faces_dir, exist_ok=True)
    known_encodings = []
    saved_count = 0
    print("🧑‍🦰 Извлекаем лучшие лица...")
    for idx, file in enumerate(sorted(os.listdir(frames_dir))):
        if not file.lower().endswith(".jpg"):
            continue
        frame_path = os.path.join(frames_dir, file)
        image = face_recognition.load_image_file(frame_path)
        face_locations = face_recognition.face_locations(image)
        face_encodings = face_recognition.face_encodings(image, face_locations)
        landmarks_list = face_recognition.face_landmarks(image)
        if not face_locations:
            print(f"Кадр {idx+1}: лиц не найдено")
            continue
        best_candidate = None
        best_score = -1
        for i, (location, encoding) in enumerate(zip(face_locations, face_encodings)):
            matches = face_recognition.compare_faces(known_encodings, encoding, tolerance=tolerance)
            if any(matches):
                continue
            top, right, bottom, left = location
            height = bottom - top
            width = right - left
            if height < 50 or width < 50:
                continue
            pad_y = int(0.3 * height)
            pad_x = int(0.2 * width)
            extended_top = max(0, top - pad_y)
            extended_bottom = min(image.shape[0], bottom + pad_y)
            extended_left = max(0, left - pad_x)
            extended_right = min(image.shape[1], right + pad_x)
            face_crop = image[extended_top:extended_bottom, extended_left:extended_right]
            sharpness = estimate_sharpness(face_crop)
            if sharpness < 100:
                continue
            is_frontal = is_frontal_face(landmarks_list[i]) if i < len(landmarks_list) else True
            frontal_score = 1.0 if is_frontal else 0.5
            area = height * width
            score = frontal_score * area * sharpness
            if score > best_score:
                best_score = score
                best_candidate = {
                    "crop": face_crop,
                    "encoding": encoding
                }
        if best_candidate:
            known_encodings.append(best_candidate["encoding"])
            pil_image = Image.fromarray(best_candidate["crop"])
            face_path = os.path.join(faces_dir, f"face_{saved_count:03d}.jpg")
            pil_image.save(face_path)
            saved_count += 1
            print(f"✅ Сохранено лицо: {face_path}")
        else:
            print(f"⏭ Кадр {idx+1}: подходящих лиц нет")
    print(f"📦 Всего уникальных лиц сохранено: {saved_count}")

extract_faces_from_frames(frames_dir, result_dir)

import os
import re
import json
import unicodedata
from collections import defaultdict
from difflib import SequenceMatcher

import numpy as np
import cv2

SIMILARITY_THRESHOLD = 0.9  # Порог похожести текста

def normalize_for_match(text):
    text = unicodedata.normalize('NFKC', text.lower())
    text = re.sub(r"[\s\n\r]+", " ", text)
    text = re.sub(r"[.,:;!?\"'’«»\\-]", "", text)
    return text.strip()

def read_image_with_unicode(path):
    try:
        stream = np.fromfile(path, dtype=np.uint8)
        img = cv2.imdecode(stream, cv2.IMREAD_COLOR)
        if img is None:
            print(f"[read_image_with_unicode] ❌ Не удалось декодировать изображение: {path}")
        else:
            print(f"[read_image_with_unicode] ✅ Загружено: {path}, размер: {img.shape}")
        return img
    except Exception as e:
        print(f"[read_image_with_unicode] ❌ Ошибка при чтении {path}: {e}")
        return None

def extract_unique_overlay_blocks_from_ocr(ocr_log_path):
    with open(ocr_log_path, encoding="utf-8") as f:
        ocr_data = json.load(f)

    grouped_blocks = {}
    frame_map = defaultdict(list)

    for fname, result in ocr_data.items():
        try:
            blocks = result["results"][0]["results"][0]["textDetection"]["pages"][0]["blocks"]
            for block in blocks:
                lines = block.get("lines", [])
                raw_text = " ".join(" ".join(word["text"] for word in line.get("words", [])) for line in lines)
                norm = normalize_for_match(raw_text)
                if not norm:
                    continue
                match_found = False
                for existing_norm in grouped_blocks:
                    if SequenceMatcher(None, norm, existing_norm).ratio() >= SIMILARITY_THRESHOLD:
                        frame_map[existing_norm].append(fname)
                        match_found = True
                        break
                if not match_found:
                    grouped_blocks[norm] = raw_text
                    frame_map[norm].append(fname)
        except Exception as e:
            print(f"[extract_unique_overlay_blocks_from_ocr]: ошибка {e}")
            continue

    return grouped_blocks, frame_map

# БЛОК АНАЛИЗА НАБИВКИ

def analyze_text_block(text_key, frames, ocr_data, frames_dir):
    overlay_counter = 0
    overlay_seconds = []
    unreadable_frames = []
    color_issues = []
    total_frames = len(frames)

    total_font_heights = 0
    total_lines_count = 0
    total_text_heights = 0
    total_text_areas = 0
    matching_frame_count = 0

    line_spacing_ratio = 1.25  # межстрочный интервал
    margin = 5  # Допуск в пикселях

    for fname in frames:
        if fname not in ocr_data:
            continue
        result = ocr_data[fname]
        try:
            blocks = result["results"][0]["results"][0]["textDetection"]["pages"][0]["blocks"]
            frame_path = os.path.join(frames_dir, fname)
            img = read_image_with_unicode(frame_path)
            if img is None:
                continue
            h_img, w_img = img.shape[:2]
            frame_area = h_img * w_img
            overlay_box = [int(w_img * 0.10), int(h_img * 0.05), int(w_img * 0.90), int(h_img * 0.95)]

            lines_this_frame = 0
            heights_this_frame = []
            line_widths = []
            found = False

            for block in blocks:
                lines = block.get("lines", [])
                block_text = normalize_for_match(" ".join(" ".join(word["text"] for word in line.get("words", [])) for line in lines))
                if SequenceMatcher(None, block_text, text_key).ratio() < SIMILARITY_THRESHOLD:
                    continue

                found = True
                lines_this_frame += len(lines)
                for line in lines:
                    line_xs = []
                    line_ys = []
                    for word in line.get("words", []):
                        vertices = word.get("boundingBox", {}).get("vertices", [])
                        try:
                            xs = [int(v["x"]) for v in vertices if "x" in v and str(v["x"]).isdigit()]
                            ys = [int(v["y"]) for v in vertices if "y" in v and str(v["y"]).isdigit()]
                        except:
                            continue
                        if len(xs) < 2 or len(ys) < 2:
                            continue
                        x_min, x_max = min(xs), max(xs)
                        y_min, y_max = min(ys), max(ys)
                        height = y_max - y_min
                        width = x_max - x_min
                        if width <= 1 or height <= 1:
                            continue
                        heights_this_frame.append(height)
                        line_xs += xs
                        line_ys += ys

                        word_crop = img[y_min:y_max, x_min:x_max]
                        if word_crop.size == 0:
                            continue

                        mean_brightness = word_crop.mean()
                        if mean_brightness > 230 or mean_brightness < 25:
                            unreadable_frames.append(fname)

                        contrast = cv2.Laplacian(cv2.cvtColor(word_crop, cv2.COLOR_BGR2GRAY), cv2.CV_64F).var()
                        if contrast < 15:
                            unreadable_frames.append(fname)

                        if not (
                            overlay_box[0] - margin <= x_min <= overlay_box[2] + margin and
                            overlay_box[0] - margin <= x_max <= overlay_box[2] + margin and
                            overlay_box[1] - margin <= y_min <= overlay_box[3] + margin and
                            overlay_box[1] - margin <= y_max <= overlay_box[3] + margin
                        ):
                            overlay_counter += 1
                            overlay_seconds.append(fname)

                        avg_color = word_crop.mean(axis=(0, 1))
                        background = img[max(y_max + 1, 0):min(y_max + 6, h_img), x_min:x_max]
                        if background.shape[0] > 0:
                            bg_color = background.mean(axis=(0, 1))
                            color_dist = np.linalg.norm(avg_color - bg_color)
                            if color_dist < 20:
                                color_issues.append(fname)

                    if line_xs:
                        line_widths.append(max(line_xs) - min(line_xs))

            if found and heights_this_frame:
                avg_font_height = sum(heights_this_frame) / len(heights_this_frame)
                avg_line_width = sum(line_widths) / len(line_widths) if line_widths else 0

                total_font_heights += avg_font_height
                total_lines_count += lines_this_frame

                text_block_height = avg_font_height * (
                    lines_this_frame + (line_spacing_ratio - 1) * max(lines_this_frame - 1, 0)
                )
                total_text_heights += text_block_height

                text_block_area = text_block_height * avg_line_width
                total_text_areas += text_block_area
                matching_frame_count += 1

        except Exception as e:
            print(f"[Ошибка анализа кадра {fname}]: {e}")
            continue

    if matching_frame_count == 0:
        return {
            "font_height_px": 0,
            "font_height_percent": 0,
            "text_block_height_percent": 0,
            "text_block_area_percent": 0,
            "lines_count": 0,
            "duration_sec": 0,
            "share_presence": 0,
            "overlay_counter": overlay_counter,
            "overlay_frames": [],
            "color_issues": [],
            "unreadable_frames": [],
        }

    # Средние метрики
    avg_font_height_px = total_font_heights / matching_frame_count
    avg_lines_count = total_lines_count / matching_frame_count
    avg_text_height = total_text_heights / matching_frame_count
    avg_text_area = total_text_areas / matching_frame_count

    font_height_percent = 100 * avg_font_height_px / h_img
    text_height_percent = 100 * avg_text_height / h_img
    text_area_percent = 100 * avg_text_area / (h_img * w_img)

    duration_sec = matching_frame_count
    total_frame_count = len([f for f in os.listdir(frames_dir) if f.endswith(".jpg")])
    share_presence = 100 * matching_frame_count / total_frame_count if total_frame_count else 0

    return {
        "font_height_px": round(avg_font_height_px, 1),
        "font_height_percent": round(font_height_percent, 1),
        "text_block_height_percent": round(text_height_percent, 2),
        "text_block_area_percent": round(text_area_percent, 2),
        "lines_count": round(avg_lines_count, 1),
        "duration_sec": duration_sec,
        "share_presence": share_presence,
        "overlay_counter": overlay_counter,
        "overlay_frames": list(set(overlay_seconds)),
        "color_issues": list(set(color_issues)),
        "unreadable_frames": list(set(unreadable_frames)),
    }

def write_overlay_report(ocr_log_path, frames_dir, output_path):
    grouped_blocks, text_map = extract_unique_overlay_blocks_from_ocr(ocr_log_path)
    with open(ocr_log_path, encoding="utf-8") as f:
        ocr_data = json.load(f)

    with open(output_path, "w", encoding="utf-8") as f:
        for norm_key, frames in text_map.items():
            raw_text = grouped_blocks[norm_key]
            print(f"\n\n🔍 Анализ набивки: {raw_text}\n-----------------------------")
            result = analyze_text_block(norm_key, frames, ocr_data, frames_dir)
            word_count = len(raw_text.split())
            reading_time = round(word_count / 2.5, 1)

            # Преобразование имен файлов в секунды
            def to_seconds_list(frame_list):
                seconds = []
                for name in frame_list:
                    try:
                        n = int(name.replace("frame_", "").replace(".jpg", ""))
                        seconds.append(f"{n}сек.")
                    except:
                        continue
                return seconds

            color_secs = to_seconds_list(result["color_issues"])
            unreadable_secs = to_seconds_list(result["unreadable_frames"])

            duration_note = ""
            if result["share_presence"] > 100:
                duration_note = " (в ролике несколько данных набивок)"

            f.write(f"ТЕКСТ НАБИВКИ: {raw_text}\n")
            f.write(f"Размер шрифта: {result['font_height_px']} px ({result['font_height_percent']}% высоты кадра)\n")
            f.write(f"Размер набивки по высоте (в среднем {result['lines_count']} строк): {result['text_block_height_percent']}% высоты кадра\n")
            f.write(f"Доля площади набивки от всего кадра: {result['text_block_area_percent']}%\n")
            f.write(f"Длительность набивки: {result['duration_sec']} секунд\n")
            f.write(f"Время на прочтение: {reading_time} секунд\n")
            f.write(f"Доля присутствия набивки: {result['share_presence']:.2f}%{duration_note}\n")
            f.write(f"Набивка не в рамке на {result['overlay_counter']} кадрах\n")

            if color_secs:
                f.write(f"Проблемы с цветом: {', '.join(color_secs)} (слишком близкий цвет текста к фону)\n")
            else:
                f.write("Проблем с цветом не обнаружено\n")

            if unreadable_secs:
                f.write(f"Плохая читаемость: {', '.join(unreadable_secs)} (низкий контраст или яркость)\n")
            else:
                f.write("Читаемость текста в норме\n")

            f.write("\n" + "-" * 60 + "\n\n")

# 📄 Извлечение текста из документов с fallback через Yandex Vision
import pdfplumber
import re
import base64
import requests

def extract_text_yandex_vision_image(filepath):
    try:
        result = recognize_image_file(
            filepath,
            api_key=yandex_api_key,
            folder_id=yandex_folder_id,
            attempts=int(os.getenv("YANDEX_VISION_RETRIES", "3")),
            timeout=int(os.getenv("YANDEX_VISION_TIMEOUT", "90")),
        )
        if result.ok:
            return result.text
        return f"[Ошибка Yandex Vision]: {result.error}"
    except Exception as e:
        return f"[Ошибка Yandex Vision]: {e}"

from pdf2image import convert_from_path
import tempfile
from pathlib import Path

from pdf2image import convert_from_path
import tempfile
from pathlib import Path

def extract_text_from_document(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    method = "неизвестный"
    try:
        if ext == ".pdf":
            text = ""
            with pdfplumber.open(filepath) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text() or ""
                    text += page_text + "\n"
            if text.strip():
                method = "pdfplumber"
                return text, method
            else:
                # Конвертируем все страницы в JPEG и отправляем в Vision
                with tempfile.TemporaryDirectory() as tmpdir:
                    images = convert_from_path(filepath, dpi=300, output_folder=tmpdir, fmt="jpeg")
                    all_text = []
                    for idx, img in enumerate(images):
                        img_path = os.path.join(tmpdir, f"page_{idx+1:03d}.jpg")
                        img.save(img_path, "JPEG")
                        vision_text = extract_text_yandex_vision_image(img_path)
                        all_text.append(vision_text)
                    method = "PDF → JPEG (все страницы) → Yandex Vision"
                    return "\n".join(all_text).strip(), method
        elif ext == ".docx":
            method = "python-docx"
            return extract_docx_text(filepath), method
        elif ext == ".txt":
            with open(filepath, "r", encoding="utf-8") as f:
                method = "txt/plain"
                return f.read(), method
        elif ext in [".jpg", ".jpeg"]:
            method = "Yandex Vision"
            return extract_text_yandex_vision_image(filepath), method
        else:
            method = "не поддерживается"
            return "[Формат не поддерживается для автоматического извлечения текста]", method
    except Exception as e:
        return f"[Ошибка извлечения текста из {filepath}: {e}]", method

# 📘 Формируем итоговый файл с текстами всех документов
documents_text_path = os.path.join(result_dir, "Documents_Texts.txt")
with open(documents_text_path, "w", encoding="utf-8") as f_out:
  counter = 1
  first = True
  for root, _, files in os.walk(documents_dir):
      for filename in sorted(files):
          file_path = os.path.join(root, filename)
          text_raw, method_used = extract_text_from_document(file_path)
          cleaned = clean_to_single_block(text_raw)
          if not first:
              f_out.write("\n\n")
          f_out.write(f"Документ {counter}. {filename}\n")
          f_out.write(f"(обработано с помощью: {method_used})\n")
          f_out.write(cleaned + "\n")
          counter += 1
          first = False

print(f"📘 Извлечённый текст из документов сохранён: {documents_text_path}")

# Путь к папке с лицами и результирующему файлу:
faces_dir = os.path.join(result_dir, "faces")
telegram_faces_output = os.path.join(result_dir, "Face_Recognition_Results.txt")

# Асинхронная функция для распознавания лиц через SearchByPhoto_Bot
async def recognize_faces_via_telegram_bot(faces_dir, result_path):
    print("✨ Начинаем распознавание лиц через Telegram-бота...")
    app = Client("face_search_bot", api_id=telegram_api_id, api_hash=telegram_api_hash)
    await app.start()

    BOT_USERNAME = "SearchByPhoto_Bot"
    bot = await app.get_users(BOT_USERNAME)
    await app.send_message(bot.id, "/start")
    await asyncio.sleep(2)

    results = []
    for filename in sorted(os.listdir(faces_dir)):
        if not filename.lower().endswith(".jpg"):
            continue

        photo_path = os.path.join(faces_dir, filename)
        print(f"📷 Обработка {filename}...")
        await app.send_photo(bot.id, photo_path)
        await asyncio.sleep(6)  # Дать время боту ответить

        # Получаем последнее сообщение от бота
        async for msg in app.get_chat_history(bot.id, limit=5):
            if msg.photo and msg.caption:
                result = msg.caption.strip()
                break
        else:
            result = "Публичная личность не установлена"

        print(f"{filename}: {result}")
        results.append(f"{filename}: {result}")
        await asyncio.sleep(3)

    with open(result_path, "w", encoding="utf-8") as f:
        for line in results:
            f.write(line + "\n")
    print(f"🔖 Результаты распознавания лиц сохранены в: {result_path}")
    await app.stop()


# 📺 Расширенный анализ эфирной версии видео: извлечение параметров

PARAM_DESCRIPTIONS = {
    "index": "Номер потока",
    "codec_name": "Название кодека",
    "codec_long_name": "Полное название кодека",
    "profile": "Профиль кодека",
    "codec_type": "Тип потока (видео, аудио)",
    "codec_tag_string": "Строковый тег кодека",
    "codec_tag": "Шестнадцатеричный тег кодека",
    "width": "Ширина кадра в пикселях",
    "height": "Высота кадра в пикселях",
    "coded_width": "Закодированная ширина",
    "coded_height": "Закодированная высота",
    "closed_captions": "Наличие скрытых субтитров",
    "film_grain": "Имитация плёнки (зернистость)",
    "has_b_frames": "Наличие B-кадров",
    "sample_aspect_ratio": "Соотношение сторон пикселя",
    "display_aspect_ratio": "Соотношение сторон изображения",
    "pix_fmt": "Формат пикселей",
    "level": "Уровень кодека",
    "color_range": "Диапазон цветов (TV/PC)",
    "color_transfer": "Тип цветопередачи",
    "color_primaries": "Цветовые примарии",
    "chroma_location": "Положение хромы",
    "field_order": "Порядок строк развёртки",
    "refs": "Число опорных кадров",
    "r_frame_rate": "Частота кадров (расчётная)",
    "avg_frame_rate": "Средняя частота кадров",
    "time_base": "Базис времени",
    "start_pts": "Начальная временная метка PTS",
    "start_time": "Начальное время (сек)",
    "duration_ts": "Продолжительность в тиках",
    "duration": "Продолжительность (сек)",
    "bit_rate": "Битрейт (бит/с)",
    "extradata_size": "Размер служебных данных кодека",
    "disposition": "Назначение дорожки (default, forced...)",
    "tags": "Метки (имя, автор, uid и пр.)",
    "side_data_list": "Дополнительные параметры кодека",
    "sample_fmt": "Формат аудиосэмплов",
    "sample_rate": "Частота дискретизации (Гц)",
    "channels": "Количество аудиоканалов",
    "bits_per_sample": "Бит на сэмпл",
    "initial_padding": "Начальный отступ аудио",
    "bits_per_raw_sample": "Бит на raw-сэмпл",
    "filename": "Имя файла",
    "nb_streams": "Количество потоков",
    "nb_programs": "Количество программ",
    "nb_stream_groups": "Количество групп потоков",
    "format_name": "Краткое имя формата",
    "format_long_name": "Полное имя формата",
    "size": "Размер файла (байт)",
    "probe_score": "Оценка достоверности анализа"
}

def describe_param(full_key):
    short_key = full_key.split(".")[-1]
    return PARAM_DESCRIPTIONS.get(short_key, "— нет описания —")

broadcast_dir = os.environ.get("BROADCAST_PATH")
param_txt = os.path.join(result_dir, "Broadcast_video_param.txt")
param_with_desc_txt = os.path.join(result_dir, "Broadcast_video_param_with_description.txt")

if broadcast_dir and os.path.exists(broadcast_dir):
    video_files = [f for f in os.listdir(broadcast_dir) if f.lower().endswith((".mxf", ".mp4"))]
    if not video_files:
        for path in [param_txt, param_with_desc_txt]:
            with open(path, "w", encoding="utf-8") as f:
                f.write("❌ В папке эфирной версии не найдено .mxf/.mp4 файлов\n")
    else:
        video_path = os.path.join(broadcast_dir, video_files[0])
        try:
            cmd = [
                "ffprobe", "-v", "quiet",
                "-show_entries", "format:stream",
                "-print_format", "json",
                video_path
            ]
            result = subprocess.run(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE, text=True)
            metadata = json.loads(result.stdout)

            plain_lines = []
            described_lines = []
            counter = 1

            for section_name in ("streams", "format"):
                entries = metadata.get(section_name, [])
                if isinstance(entries, dict):
                    entries = [entries]
                for entry in entries:
                    for key, value in entry.items():
                        full_key = f"{section_name[:-1]}.{key}"
                        line = f'{counter}. "{full_key}": "{value}"'
                        desc = describe_param(full_key)
                        line_with_desc = f'{counter}. "{desc} - " "{full_key}": "{value}"'
                        plain_lines.append(line)
                        described_lines.append(line_with_desc)
                        counter += 1

            with open(param_txt, "w", encoding="utf-8") as f:
                f.write("\n".join(plain_lines))
            with open(param_with_desc_txt, "w", encoding="utf-8") as f:
                f.write("\n".join(described_lines))

            print(f"📊 Параметры эфирной версии сохранены в:\n  - {param_txt}\n  - {param_with_desc_txt}")

        except Exception as e:
            for path in [param_txt, param_with_desc_txt]:
                with open(path, "w", encoding="utf-8") as f:
                    f.write(f"❌ Ошибка при извлечении параметров: {e}")
else:
    for path in [param_txt, param_with_desc_txt]:
        with open(path, "w", encoding="utf-8") as f:
            f.write("❌ Путь к эфирной версии не задан или не существует\n")

from PIL import Image, ImageDraw, ImageFont

def create_annotated_pdf(frames_dir, output_pdf_path):
    images = []
    frame_files = sorted([
        f for f in os.listdir(frames_dir)
        if f.lower().endswith(".jpg")
    ])

    for idx, file in enumerate(frame_files):
        path = os.path.join(frames_dir, file)
        img = Image.open(path).convert("RGB")
        draw = ImageDraw.Draw(img)

        # Размер шрифта: 5% от площади изображения
        w, h = img.size
        font_size = int((w * h) ** 0.5 * 0.05)

        # Используем стандартный шрифт (можно заменить на TTF)
        try:
            font = ImageFont.truetype("arialbd.ttf", font_size)
        except:
            font = ImageFont.load_default()

        label = f"{idx + 1}sec"
        draw.text((10, 10), label, font=font, fill=(255, 255, 255))

        images.append(img)

    if images:
        images[0].save(
            output_pdf_path,
            save_all=True,
            append_images=images[1:],
            resolution=100
        )
        print(f"📄 PDF с подписями создан: {output_pdf_path}")
    else:
        print("❌ Нет кадров для PDF")

# 📍 Генерация PDF с аннотациями
annotated_pdf_path = os.path.join(result_dir, "Annotated_Frames.pdf")
create_annotated_pdf(frames_dir, annotated_pdf_path)

def create_faces_pdf(faces_dir, output_pdf_path):
    images = []
    face_files = sorted([
        f for f in os.listdir(faces_dir)
        if f.lower().endswith(".jpg")
    ])

    for idx, file in enumerate(face_files):
        path = os.path.join(faces_dir, file)
        img = Image.open(path).convert("RGB")
        draw = ImageDraw.Draw(img)

        # Уменьшаем изображение до ширины 400px
        w_target = 400
        w, h = img.size
        if w > w_target:
            ratio = w_target / w
            img = img.resize((w_target, int(h * ratio)))

        # Подпись
        try:
            font = ImageFont.truetype("arialbd.ttf", 16)
        except:
            font = ImageFont.load_default()

        draw.text((10, 10), f"{idx+1}", font=font, fill=(255, 255, 255))
        images.append(img)

    if images:
        images[0].save(output_pdf_path, save_all=True, append_images=images[1:])
        print(f"📄 PDF из лиц создан: {output_pdf_path}")
    else:
        print("⚠️ Лица не найдены для PDF")

# Вызов
faces_pdf_path = os.path.join(result_dir, "Faces_Thumbnails.pdf")
create_faces_pdf(faces_dir, faces_pdf_path)

# 📋 Проверка параметров на соответствие техническим требованиям
def check_tech_requirements(param_file_path, output_json_path):
    REQUIRED_PARAMS = {
        "forma.format_name": ("mxf", "MXF"),
        "forma.format_long_name": ("MXF (Material eXchange Format)", "MXF (Material eXchange Format)"),
        "forma.tags.operational_pattern_ul": ("060e2b34.04010101.0d010201.01010900", "OP-1a"),
        "stream.codec_name": ("mpeg2video", "MPEG Video"),
        "stream.codec_long_name": ("MPEG-2 video", "XDCAM HD422"),
        "stream.profile": ("4:2:2", "4:2:2@High"),
        "stream.r_frame_rate": ("25/1", "25 кадров/сек"),
        "stream.avg_frame_rate": ("25/1", "25 кадров/сек"),
        "stream.bit_rate": (">=50000000", "≥ 50 Мбит/с"),
        "stream.pix_fmt": ("yuv422p", "4:2:2"),
        "stream.width": ("1920", "1920 пикселей"),
        "stream.height": ("1080", "1080 пикселей"),
        "stream.display_aspect_ratio": ("16:9", "16:9"),
        "stream.sample_aspect_ratio": ("1:1", "SQUARE 1:1"),
        "stream.color_primaries": ("bt709", "YUV (bt709)"),
        "stream.color_transfer": ("bt709", "YUV (bt709)"),
        "stream.field_order": ("tt", "Чересстрочная развёртка"),
        "stream.chroma_location": ("topleft", "Верхнее поле первое"),
        "stream.start_time": ("0.000000", "00:00:00:00"),
        "forma.bit_rate": (">=50000000", "≥ 50 Мбит/с"),
        "audio.stream.codec_name": ("pcm_s24le", "PCM"),
        "audio.stream.sample_fmt": ("s32", "16 / 24 бит"),
        "audio.stream.bit_rate": (">=768000", "768 - 1152 Кбит/с"),
        "audio.stream.sample_rate": ("48000", "48000 Гц"),
        "audio.stream.channels": ("1", "1 канал на 1 трек"),
    }

    results = []
    try:
        with open(param_file_path, encoding="utf-8") as f:
            lines = f.readlines()

        for param_key, (expected_val, display_range) in REQUIRED_PARAMS.items():
            actual_val = None
            for line in lines:
                if f'"{param_key}": "' in line:
                    actual_val = line.split(f'"{param_key}": "')[-1].strip().strip('"')
                    break

            if actual_val is None:
                results.append({
                    "param": param_key,
                    "actual": "[не найдено]",
                    "expected": expected_val,
                    "range": display_range,
                    "ok": False,
                    "not_supported": False
                })
                continue

            ok = False
            if expected_val.startswith(">="):
                try:
                    ok = int(actual_val) >= int(expected_val.replace(">=", ""))
                except:
                    ok = False
            else:
                ok = actual_val.lower() == expected_val.lower()

            results.append({
                "param": param_key,
                "actual": actual_val,
                "expected": expected_val,
                "range": display_range,
                "ok": ok,
                "not_supported": False
            })

        # Добавим оставшиеся 10, которые не проверяются автоматически
        not_supported = [
            "forma.format_version", "stream.gop_structure", "stream.bitrate_mode",
            "stream.standard", "stream.bit_depth", "stream.version_tag", "stream.video_id",
            "audio.stream.bitrate_mode", "audio.stream.stereo_mode", "audio.stream.track_ids"
        ]
        for p in not_supported:
            results.append({
                "param": p,
                "actual": "[не проверяется]",
                "expected": "-",
                "range": "-",
                "ok": False,
                "not_supported": True
            })

        with open(output_json_path, "w", encoding="utf-8") as f:
            json.dump(results, f, ensure_ascii=False, indent=2)
        print(f"📋 Результаты проверки сохранены: {output_json_path}")
    except Exception as e:
        with open(output_json_path, "w", encoding="utf-8") as f:
            json.dump([{"param": "Ошибка", "actual": "", "expected": "", "range": "", "ok": False, "message": str(e)}], f)

# Вызов функции:
tech_validation_json = os.path.join(result_dir, "Tech_Validation_Result.json")
check_tech_requirements(param_with_desc_txt, tech_validation_json)

# 🆕 Человекочитаемая выжимка техпроверки для GPT
broadcast_check_txt = os.path.join(result_dir, "broadcast_check.txt")
try:
    with open(tech_validation_json, "r", encoding="utf-8") as f:
        results = json.load(f)

    ok_cnt = fail_cnt = na_cnt = 0
    lines = []
    lines.append("Итоговая проверка эфирного видео (broadcast)")
    lines.append("-" * 60)

    for r in results:
        # not_supported → это пункты, которые автоматически не проверяются
        if r.get("not_supported"):
            na_cnt += 1
            status = "N/A"
        else:
            status = "OK" if r.get("ok") else "FAIL"
            if status == "OK":
                ok_cnt += 1
            else:
                fail_cnt += 1

        param = r.get("param", "—")
        actual = r.get("actual", "—")
        expected = r.get("expected", "—")
        display_range = r.get("range", "—")
        lines.append(f"{param}: {actual} | ожидается: {expected} ({display_range}) => {status}")

    lines.append("-" * 60)
    lines.append(f"Сводка: OK={ok_cnt} | FAIL={fail_cnt} | N/A={na_cnt}")

    with open(broadcast_check_txt, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    print(f"📑 Итоговый файл проверки эфирного видео сохранён: {broadcast_check_txt}")
except Exception as e:
    with open(broadcast_check_txt, "w", encoding="utf-8") as f:
        f.write(f"❌ Ошибка при формировании broadcast_check.txt: {e}\n")
    print(f"❌ Ошибка при формировании broadcast_check.txt: {e}")


EXCEL_SOURCE_DIR = r"C:\Users\PTambulatov\Desktop\PT\2.Расчеты\50. ХАКАТОН\11. Опыт баера"
buyer_experience_txt = os.path.join(result_dir, "Buyer_Experience.txt")

BUYER_EXPERIENCE_PATH = os.path.join(EXCEL_SOURCE_DIR, "Buyer_Experience.xlsx")
BUYER_EXPERIENCE_JSON_PATH = os.path.join(result_dir, "Buyer_Experience.json")

def build_buyer_experience_json(excel_path_or_dir: str,
                                out_json_path: str | None = None) -> str:
    """
    Конвертирует Buyer_Experience.xlsx в структурированный JSON формата:
    {
      "version": "YYYY-MM-DD",
      "source": "Buyer_Experience.xlsx",
      "generated_at": "ISO-8601 UTC",
      "fields": {
        "Категория": "category",
        "Что проверяем": "check",
        "Требование законодательства": "requirement"
      },
      "checks": [ { id, category, check, requirement, legal_ref, severity, ... } ]
    }
    Возвращает путь к созданному JSON. При ошибке всё равно создаёт минимальный JSON с полем "error".
    """
    import os, json, re
    from datetime import datetime, timezone

    # 1) Определяем входной Excel
    if os.path.isdir(excel_path_or_dir):
        # найдём первый .xlsx/.xlsm
        excel_files = sorted([f for f in os.listdir(excel_path_or_dir)
                              if f.lower().endswith((".xlsx", ".xlsm"))])
        if not excel_files:
            payload = {"error": "excel_not_found_in_dir", "dir": excel_path_or_dir}
            outp = out_json_path or os.path.join(excel_path_or_dir, "Buyer_Experience.json")
            os.makedirs(os.path.dirname(outp), exist_ok=True)
            with open(outp, "w", encoding="utf-8") as f:
                json.dump(payload, f, ensure_ascii=False, indent=2)
            return outp
        excel_path = os.path.join(excel_path_or_dir, excel_files[0])
    else:
        excel_path = excel_path_or_dir

    folder = os.path.dirname(excel_path) or "."
    outp = out_json_path or os.path.join(folder, "Buyer_Experience.json")
    os.makedirs(os.path.dirname(outp), exist_ok=True)

    def _norm(s: str) -> str:
        return re.sub(r"\s+", " ", str(s or "")).strip().lower()

    # 2) Синонимы заголовков → канонические ключи
    SYN = {
        "category": [
            "категория", "категория (блок)", "блок", "раздел"
        ],
        "check": [
            "что проверяем", "проверка", "критерий", "пункт", "описание проверки"
        ],
        "requirement": [
            "требование законодательства", "требование", "норма", "описание требования"
        ],
        "legal_ref": [
            "ссылка на норму", "нормативная ссылка", "норма закона", "правовая ссылка", "статья"
        ],
        "severity": [
            "критичность", "уровень риска", "уровень", "важность", "severity", "приоритет"
        ],
        "min_duration_sec": [
            "минимальная длительность (сек)", "мин длительность", "длительность", "min_duration_sec", "минимальная длительность"
        ],
        "min_area_percent": [
            "минимальная площадь (%)", "мин площадь", "% площади", "min_area_percent", "минимальный размер (%)"
        ],
        "notes": [
            "примечания", "пример", "комментарий", "замечания", "notes"
        ],
    }

    def _severity_norm(v: str) -> str | None:
        t = _norm(v)
        if not t:
            return None
        mapping = {
            "высокий": ["high", "выс", "h", "критичный"],
            "средний": ["medium", "сред", "m"],
            "низкий":  ["low", "низ", "l", "несущественный"],
        }
        for k, aliases in mapping.items():
            if t == k or any(t == a for a in aliases):
                return k
        return t  # оставим как есть

    def _to_int(v):
        if v is None:
            return None
        if isinstance(v, (int, float)):
            try:
                return int(v)
            except Exception:
                return None
        m = re.search(r"-?\d+", str(v))
        return int(m.group(0)) if m else None

    def _to_percent(v):
        if v is None:
            return None
        s = str(v).strip().replace(",", ".")
        m = re.search(r"-?\d+(\.\d+)?", s)
        if not m:
            return None
        try:
            val = float(m.group(0))
            # 7  или 7.0 трактуем как проценты «как есть»
            return val
        except Exception:
            return None

    # 3) Чтение Excel — pandas → openpyxl
    frames = []
    fields_map_agg = {}  # "Оригинальное имя столбца" -> canonical
    try:
        try:
            import pandas as pd  # type: ignore
            xl = pd.ExcelFile(excel_path)
            for sheet_name in xl.sheet_names:
                df = xl.parse(sheet_name=sheet_name, dtype=str)
                if df.empty:
                    continue
                cols = [str(c) for c in df.columns]
                # Построим отображение колонок
                col_to_canon = {}
                for i, col in enumerate(cols):
                    n = _norm(col)
                    for canon, syns in SYN.items():
                        if n in [_norm(x) for x in syns]:
                            col_to_canon[i] = canon
                            # для "fields" запомним первое попадание красивым именем
                            fields_map_agg.setdefault(col, canon)
                            break

                # Идём по строкам
                for ridx, row in df.iterrows():
                    get = lambda canon: row[cols[i]] if any(i for i,c in col_to_canon.items() if c==canon) else None
                    item = {}
                    # Базовые поля
                    if any(canon in col_to_canon.values() for canon in ("category","check","requirement","legal_ref","severity","min_duration_sec","min_area_percent","notes")):
                        for i, col in enumerate(cols):
                            canon = col_to_canon.get(i)
                            if not canon:
                                continue
                            val = row[col]
                            if pd.isna(val):
                                val = None
                            if canon == "severity":
                                val = _severity_norm(val)
                            elif canon == "min_duration_sec":
                                val = _to_int(val)
                            elif canon == "min_area_percent":
                                val = _to_percent(val)
                            item[canon] = val if (val is None or str(val).strip() != "") else None
                        frames.append(item)
        except Exception:
            # openpyxl fallback (более общий, но без «умных» типов)
            from openpyxl import load_workbook  # type: ignore
            wb = load_workbook(excel_path, data_only=True, read_only=True)
            for ws in wb.worksheets:
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                # найдём хедер — первая строка с 2+ непустыми ячейками
                header = None
                header_idx = 0
                for i, r in enumerate(rows[:20]):
                    non_empty = [x for x in r if x not in (None, "")]
                    if len(non_empty) >= 2:
                        header = [str(x or "") for x in r]
                        header_idx = i
                        break
                if not header:
                    continue
                col_to_canon = {}
                for i, col in enumerate(header):
                    n = _norm(col)
                    for canon, syns in SYN.items():
                        if n in [_norm(x) for x in syns]:
                            col_to_canon[i] = canon
                            fields_map_agg.setdefault(col, canon)
                            break
                for r in rows[header_idx+1:]:
                    item = {}
                    for i, val in enumerate(list(r)):
                        canon = col_to_canon.get(i)
                        if not canon:
                            continue
                        if val is None:
                            v = None
                        else:
                            v = str(val)
                        if canon == "severity":
                            v = _severity_norm(v)
                        elif canon == "min_duration_sec":
                            v = _to_int(v)
                        elif canon == "min_area_percent":
                            v = _to_percent(v)
                        item[canon] = v if (v is None or str(v).strip() != "") else None
                    if item:
                        frames.append(item)
    except Exception as e:
        with open(outp, "w", encoding="utf-8") as f:
            json.dump({"error": f"excel_read_failed: {e}", "source": os.path.basename(excel_path)}, f, ensure_ascii=False, indent=2)
        return outp

    # 4) Нормализация/нумерация и сбор итогового JSON
    checks = []
    idx = 1
    for it in frames:
        # пропустим полностью пустые строки
        if not any(v not in (None, "", []) for v in it.values()):
            continue
        rec = {"id": idx}
        rec.update(it)
        checks.append(rec)
        idx += 1

    # Заполним обязательные верхние поля
    try:
        mtime = os.path.getmtime(excel_path)
        version = datetime.fromtimestamp(mtime).strftime("%Y-%m-%d")
    except Exception:
        version = datetime.utcnow().strftime("%Y-%m-%d")

    payload = {
        "version": version,
        "source": os.path.basename(excel_path),
        "generated_at": datetime.utcnow().replace(tzinfo=timezone.utc).isoformat().replace("+00:00", "Z"),
        "fields": {}
    }
    # В "fields" выводим только три ключевых отображения, если распознали
    for pretty, canon in list(fields_map_agg.items()):
        if canon in ("category", "check", "requirement"):
            payload["fields"][pretty] = canon

    payload["checks"] = checks

    try:
        with open(outp, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        print(f"🧾 Buyer_Experience.json создан: {outp} (элементов: {len(checks)})")
    except Exception as e:
        with open(outp, "w", encoding="utf-8") as f:
            json.dump({"error": f"json_write_failed: {e}"}, f, ensure_ascii=False, indent=2)

    return outp


def build_buyer_experience_from_excel(excel_dir, output_path):
    try:
        if not os.path.exists(excel_dir):
            msg = f"❌ Папка не найдена: {excel_dir}"
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(msg)
            print(msg)
            return

        excel_files = sorted([
            f for f in os.listdir(excel_dir)
            if f.lower().endswith((".xlsx", ".xlsm"))
        ])
        if not excel_files:
            msg = "❌ В папке не найдено Excel-файлов (.xlsx/.xlsm)"
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(msg)
            print(msg)
            return

        excel_path = os.path.join(excel_dir, excel_files[0])

        # Импортируем локально, чтобы не ломать запуск, если пакета нет
        try:
            import openpyxl
        except Exception as e:
            msg = f"❌ Не установлен пакет openpyxl: {e}"
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(msg)
            print(msg)
            return

        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active

        def cell_to_text(val):
            if val is None:
                return ""
            s = str(val).replace("\r", " ").replace("\n", " ").strip()
            return s

        lines = []
        idx = 1
        # Идём по всем строкам листа, берём колонки D (4) и E (5)
        for d_val, e_val in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=4, max_col=5, values_only=True):
            d_text = cell_to_text(d_val)
            e_text = cell_to_text(e_val)
            if d_text:  # Только непустые D
                lines.append(f"{idx}. {d_text} : {e_text}")
                idx += 1

        content = "\n".join(lines) if lines else "ℹ️ В столбце D нет непустых строк."
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(content)

        wb.close()
        print(f"📊 Сводка из Excel сохранена: {output_path}\n   Источник: {excel_path}")
    except Exception as e:
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(f"❌ Ошибка обработки Excel: {e}")
        print(f"❌ Ошибка обработки Excel: {e}")

# Вызов формирования отчёта из Excel
build_buyer_experience_from_excel(EXCEL_SOURCE_DIR, buyer_experience_txt)

# 🆕 Дополнительно: конвертация Excel -> JSON (в ту же папку, что и Excel)
try:
    build_buyer_experience_json(BUYER_EXPERIENCE_PATH, BUYER_EXPERIENCE_JSON_PATH)
except Exception as e:
    print(f"⚠️ Не удалось сгенерировать Buyer_Experience.json: {e}")


# =============== НОВЫЙ БЛОК: YANDEXGPT (с логированием размеров «загружаемых» файлов) ===============
import os, json, re, requests

# env/дефолты
try:
    yandex_api_key
except NameError:
    yandex_api_key = os.getenv("YANDEX_API_KEY", "")

try:
    yandex_folder_id
except NameError:
    yandex_folder_id = os.getenv("YANDEX_FOLDER_ID", "")

# Модель YandexGPT (можно переключить на "yandexgpt")
YAGPT_MODEL_NAME = os.getenv("CLIPCHECKER_YAGPT_MODEL", "yandexgpt-lite")  # "yandexgpt-lite" | "yandexgpt"
YAGPT_TEMPERATURE = 0.2
YAGPT_MAX_TOKENS  = 1500

# Строгий путь к общему промпту
PROMPT_GENERAL_FILE = r"C:\Users\PTambulatov\Desktop\PT\2.Расчеты\50. ХАКАТОН\12. Prompt\1.Prompt_general.txt"

# Ограничения на размерах
MAX_TEXT_PER_SECTION = 120_000   # символов на раздел
MAX_TOTAL_CONTEXT    = 280_000   # общий лимит контекста
MAX_FRAMES_PICK      = 60        # не более 60 (каждый 3-й)
FRAME_STEP           = 3
MAX_FACES_LIST       = 200

def _read_text_file_smart(path: str) -> str:
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            with open(path, "r", encoding=enc, errors="strict") as f:
                return f.read()
        except FileNotFoundError:
            return ""
        except UnicodeDecodeError:
            continue
        except Exception:
            continue
    return ""

def _find_by_basename(result_dir: str, base: str, with_txt_optional=True) -> str:
    cand = []
    try:
        for name in os.listdir(result_dir):
            low = name.lower()
            if with_txt_optional:
                if low == base.lower() or low == (base + ".txt").lower():
                    cand.append(os.path.join(result_dir, name))
            else:
                if low == base.lower():
                    cand.append(os.path.join(result_dir, name))
    except Exception:
        return ""
    return cand[0] if cand else ""

def _safe_getsize(p: str) -> int:
    try:
        return os.path.getsize(p)
    except Exception:
        return 0

def _collect_materials_for_yagpt(result_dir: str) -> dict:
    """
    Собирает материалы для YandexGPT.
    Возвращает:
      frames: [paths]
      faces:  [paths]
      texts:  {section: content}
      text_paths: {section: path or ""}  ← для логирования размеров
      summary: {...}
    """
    frames_dir = os.path.join(result_dir, "frames")
    faces_dir  = os.path.join(result_dir, "faces")

    # каждый 3-й кадр
    frames = []
    try:
        all_frames = sorted([f for f in os.listdir(frames_dir) if f.lower().endswith((".jpg",".jpeg",".png",".webp"))])
        for i in range(0, len(all_frames), max(1, FRAME_STEP)):
            frames.append(os.path.join(frames_dir, all_frames[i]))
        if len(frames) > MAX_FRAMES_PICK:
            frames = frames[:MAX_FRAMES_PICK]
    except Exception:
        pass

    # лица
    faces = []
    try:
        all_faces = sorted([f for f in os.listdir(faces_dir) if f.lower().endswith((".jpg",".jpeg",".png",".webp"))])
        faces = [os.path.join(faces_dir, f) for f in all_faces[:MAX_FACES_LIST]]
    except Exception:
        pass

    # текстовые файлы (пути + содержание)
    candidates = {
        "All_Frames_Text": _find_by_basename(result_dir, "All_Frames_Text"),
        "Documents_Texts.txt": _find_by_basename(result_dir, "Documents_Texts.txt", with_txt_optional=False),
        "Music.txt": _find_by_basename(result_dir, "Music.txt", with_txt_optional=False),
        "Music_Shazam.txt": _find_by_basename(result_dir, "Music_Shazam.txt", with_txt_optional=False),
        "Overlay_Report.txt": _find_by_basename(result_dir, "Overlay_Report.txt", with_txt_optional=False),
        "Speech_Yandex.txt": _find_by_basename(result_dir, "Speech_Yandex.txt", with_txt_optional=False),
    }

    texts, text_paths = {}, {}
    for key, path in candidates.items():
        text_paths[key] = path or ""
        if path and os.path.isfile(path):
            raw = _read_text_file_smart(path)
            texts[key] = (raw[:MAX_TEXT_PER_SECTION] + "\n\n…(обрезано)…") if len(raw) > MAX_TEXT_PER_SECTION else raw
        else:
            texts[key] = ""

    # Buyer experience (вне result_dir)
    buyer_text, buyer_meta = _read_buyer_experience_text(BUYER_EXPERIENCE_PATH)
    texts["Buyer_Experience"] = buyer_text
    text_paths["Buyer_Experience"] = buyer_meta.get("path", BUYER_EXPERIENCE_PATH)

    summary = {
        "frames_count": len(frames),
        "faces_count": len(faces),
        "texts_present": [k for k, v in texts.items() if v],
        "texts_missing": [k for k, v in texts.items() if not v],
        "buyer_experience_found": buyer_meta.get("found", False),
        "buyer_experience_chars": len(buyer_text or ""),
        "buyer_experience_path": buyer_meta.get("path", BUYER_EXPERIENCE_PATH),
        "buyer_experience_error": buyer_meta.get("error"),
    }
    return {"frames": frames, "faces": faces, "texts": texts, "text_paths": text_paths, "summary": summary}

def _read_buyer_experience_text(path: str) -> tuple[str, dict]:
    meta = {"path": path, "found": False, "error": None}
    if not os.path.isfile(path):
        meta["error"] = "file_not_found"
        return "", meta

    # Сначала попробуем pandas
    try:
        import pandas as pd  # type: ignore
        try:
            sheets = pd.read_excel(path, sheet_name=None, dtype=str)
            buff = []
            for sname, df in sheets.items():
                buff.append(f"Лист: {sname}")
                df = df.iloc[:200, :30].fillna("")
                for row in df.itertuples(index=False, name=None):
                    buff.append("\t".join(str(x) for x in row))
                buff.append("")
            text = "\n".join(buff).strip()
            if len(text) > MAX_TEXT_PER_SECTION:
                text = text[:MAX_TEXT_PER_SECTION] + "\n\n…(обрезано)…"
            meta["found"] = True
            return text, meta
        except Exception as e_pd:
            meta["error"] = f"pandas_failed: {e_pd}"
    except Exception:
        meta["error"] = "pandas_not_available"

    # Fallback на openpyxl
    try:
        from openpyxl import load_workbook  # type: ignore
        try:
            wb = load_workbook(path, data_only=True, read_only=True)
            buff = []
            for ws in wb.worksheets:
                buff.append(f"Лист: {ws.title}")
                for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                    vals = [("" if v is None else str(v)) for v in list(row)[:30]]
                    buff.append("\t".join(vals))
                    if i >= 200: break
                buff.append("")
            text = "\n".join(buff).strip()
            if len(text) > MAX_TEXT_PER_SECTION:
                text = text[:MAX_TEXT_PER_SECTION] + "\n\n…(обрезано)…"
            meta["found"] = True
            return text, meta
        except Exception as e_ox:
            meta["error"] = f"openpyxl_failed: {e_ox}"
    except Exception:
        if meta["error"] is None:
            meta["error"] = "openpyxl_not_available"

    return "", meta

def _build_context_for_yagpt(materials: dict) -> str:
    frames = materials.get("frames", [])
    faces  = materials.get("faces", [])
    texts  = materials.get("texts", {})
    summary = materials.get("summary", {})

    parts = []
    parts.append("=== КРАТКАЯ СВОДКА МАТЕРИАЛОВ ===")
    parts.append(f"Кадры (frames): {summary.get('frames_count',0)} шт. (каждый {FRAME_STEP}-й).")
    parts.append(f"Лица (faces): {summary.get('faces_count',0)} шт.")
    if summary.get("texts_present"):
        parts.append("Текстовые разделы: " + ", ".join(summary["texts_present"]))
    if summary.get("texts_missing"):
        parts.append("Отсутствуют: " + ", ".join(summary["texts_missing"]))
    if summary.get("buyer_experience_found"):
        parts.append("Файл Buyer_Experience: найден и включён.")
    else:
        parts.append(f"Файл Buyer_Experience: не найден/не прочитан ({summary.get('buyer_experience_error')}).")
    parts.append("")

    if frames:
        parts.append("=== СПИСОК ФАЙЛОВ: FRAMES (каждый 3-й) ===")
        for p in frames[:200]: parts.append("• " + p)
        if len(frames) > 200: parts.append(f"… и ещё {len(frames)-200} файлов")
        parts.append("")

    if faces:
        parts.append("=== СПИСОК ФАЙЛОВ: FACES ===")
        for p in faces[:200]: parts.append("• " + p)
        if len(faces) > 200: parts.append(f"… и ещё {len(faces)-200} файлов")
        parts.append("")

    for title, content in [
        ("=== ТЕКСТ С КАДРОВ (All_Frames_Text) ===", texts.get("All_Frames_Text","")),
        ("=== ТЕКСТЫ ДОКУМЕНТОВ (Documents_Texts.txt) ===", texts.get("Documents_Texts.txt","")),
        ("=== МУЗЫКА (Music.txt) ===", texts.get("Music.txt","")),
        ("=== МУЗЫКА SHAZAM (Music_Shazam.txt) ===", texts.get("Music_Shazam.txt","")),
        ("=== ОТЧЁТ ПО НАБИВКАМ (Overlay_Report.txt) ===", texts.get("Overlay_Report.txt","")),
        ("=== ОЗВУЧКА (Speech_Yandex.txt) ===", texts.get("Speech_Yandex.txt","")),
        ("=== ОПЫТ БАЙЕРА (Buyer_Experience) ===", texts.get("Buyer_Experience","")),
    ]:
        if content:
            parts.append(title); parts.append(content); parts.append("")

    context = "\n".join(parts)
    if len(context) > MAX_TOTAL_CONTEXT:
        context = context[:MAX_TOTAL_CONTEXT] + "\n\n…(контекст укорочен)…"
    return context

def _read_prompt_general_strict() -> str:
    path = os.path.normpath(PROMPT_GENERAL_FILE)
    parent = os.path.dirname(path); base = os.path.basename(path)
    print(f"🔎 Проверка промпта (general): {path}")
    print(f"  exists={os.path.exists(path)}  isfile={os.path.isfile(path)}")

    cand = path
    if not os.path.isfile(cand):
        try:
            for name in os.listdir(parent):
                if name.lower() == base.lower():
                    cand = os.path.join(parent, name); break
        except Exception as e:
            print(f"  ⚠️ Не удалось прочитать каталог: {e}")

    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            with open(cand, "r", encoding=enc, errors="strict") as f:
                t = f.read().strip()
            print(f"🧾 Читаем промпт (general) из файла ({enc}): {cand}")
            return t
        except Exception:
            continue
    print("❌ Не удалось прочитать общий промпт.")
    return ""

def _extract_text_from_yagpt_response(result_obj) -> str:
    try:
        if isinstance(result_obj, dict):
            root = result_obj.get("result") or result_obj
            alts = root.get("alternatives")
            if isinstance(alts, list) and alts:
                msg = alts[0].get("message", {})
                if isinstance(msg, dict):
                    t = msg.get("text")
                    if isinstance(t, str): return t.strip()
            if "text" in root and isinstance(root["text"], str):
                return root["text"].strip()
            return json.dumps(result_obj, ensure_ascii=False)
        return str(result_obj)
    except Exception:
        return str(result_obj)

def _yagpt_complete_with_sdk(sdk, model_uri: str, messages: list, temperature: float, max_tokens: int):
    model = sdk.models.completions(model_uri).configure(
        temperature=temperature, max_tokens=max_tokens
    )
    last_err = None
    for method in ("run", "chat", "infer"):
        if hasattr(model, method):
            try:
                res = getattr(model, method)(messages=messages)
                return res, f"sdk.{method}"
            except TypeError:
                try:
                    res = getattr(model, method)(request={"messages": messages})
                    return res, f"sdk.{method}(request=...)"
                except Exception as e2:
                    last_err = e2
            except Exception as e:
                last_err = e
    raise AttributeError(f"No suitable SDK method; last error: {last_err}")

def _yagpt_complete_with_rest(model_uri: str, messages: list, temperature: float, max_tokens: int):
    url = "https://llm.api.cloud.yandex.net/foundationModels/v1/completion"
    headers = {
        "Authorization": f"Api-Key {yandex_api_key}",
        "x-folder-id": yandex_folder_id,
        "Content-Type": "application/json",
    }
    payload = {
        "modelUri": model_uri,
        "completionOptions": {
            "stream": False,
            "temperature": temperature,
            "maxTokens": max_tokens,
        },
        "messages": messages,
    }
    r = requests.post(url, headers=headers, json=payload, timeout=90)
    r.raise_for_status()
    return r.json(), "rest.completion"

def run_yagpt(result_dir: str) -> str:
    """
    Собираем материалы из Clipchecker_materials (+ Buyer_Experience.xlsx) и отдаём их в YandexGPT.
    Выход: Comprehensive_Legal_Review.txt (для правой колонки Flask),
           лог — VLM_Connection_Log.json (совместимость по имени) с размерами файлов.
    """
    print("\n========================")
    print("🧠 Запуск анализа в YandexGPT (текстовая сводка всех материалов + опыт байера)")
    print("========================")

    # Пути вывода
    review_path     = os.path.join(result_dir, "Comprehensive_Legal_Review.txt")
    conn_log_path   = os.path.join(result_dir, "VLM_Connection_Log.json")  # совместимость по имени
    model_used_path = os.path.join(result_dir, "VLM_Model_Used.txt")

    # 1) Сбор материалов
    materials = _collect_materials_for_yagpt(result_dir)

    # 1a) Подсчёт размеров файлов
    frames_sizes = [{"name": os.path.basename(p), "bytes": _safe_getsize(p)} for p in materials["frames"]]
    faces_sizes  = [{"name": os.path.basename(p), "bytes": _safe_getsize(p)} for p in materials["faces"]]
    text_file_sizes = {}
    total_text_bytes = 0
    for key, p in materials.get("text_paths", {}).items():
        b = _safe_getsize(p) if p else 0
        included_chars = len(materials["texts"].get(key, "") or "")
        text_file_sizes[key] = {"path": p or "", "file_bytes": b, "included_chars": included_chars}
        total_text_bytes += b

    buyer_bytes = _safe_getsize(materials["summary"].get("buyer_experience_path", "")) if materials["summary"].get("buyer_experience_found") else 0

    total_frames_bytes = sum(item["bytes"] for item in frames_sizes)
    total_faces_bytes  = sum(item["bytes"] for item in faces_sizes)
    total_all_bytes    = total_frames_bytes + total_faces_bytes + total_text_bytes + buyer_bytes

    print(f"📦 Размеры файлов — frames: {total_frames_bytes} B ({len(frames_sizes)} шт.)")
    print(f"📦 Размеры файлов — faces : {total_faces_bytes} B ({len(faces_sizes)} шт.)")
    print(f"📦 Размеры файлов — texts : {total_text_bytes} B ({len(text_file_sizes)} секций) + buyer={buyer_bytes} B")
    print(f"📦 ИТОГО по материалам: {total_all_bytes} B")

    # 2) Промпт
    prompt_general = _read_prompt_general_strict()
    if not prompt_general:
        msg = ("❌ Не удалось прочитать текст промпта из 1.Prompt_general.txt по строгому пути.\n"
               f"Ожидался файл: {PROMPT_GENERAL_FILE}")
        with open(review_path, "w", encoding="utf-8") as f: f.write(msg)
        with open(model_used_path, "w", encoding="utf-8") as f: f.write("—")
        with open(conn_log_path, "w", encoding="utf-8") as f:
            json.dump({"error": "prompt_general_not_found", "path": PROMPT_GENERAL_FILE}, f, ensure_ascii=False, indent=2)
        print(msg); return ""

    # 3) Контекст
    context_text = _build_context_for_yagpt(materials)
    payload_preview_len = min(1000, len(context_text))

    # 4) Сообщения
    SYSTEM_TEXT = (
        "Ты делаешь предварительный скрининг рекламных материалов для ТВ. Отвечай по-русски, по чек-листу, "
        "без дисклеймеров и повторов. Запрещено использовать формулировки вроде «Недостаточно данных», "
        "«требуется полная информация» или просьбы прислать ещё документы. "
        "Если каких-то материалов не хватает — всё равно делай выводы по имеющимся данным и "
        "в конце добавь отдельный раздел «Чего не хватает» (списком). "
        "Форматируй вывод компактно, с подзаголовками."
    )
    messages = [
        {"role": "system", "text": SYSTEM_TEXT},
        {"role": "user",   "text": f"{prompt_general}\n\n---\nМАТЕРИАЛЫ ДЛЯ АНАЛИЗА:\n{context_text}"}
    ]

    model_uri = f"gpt://{yandex_folder_id}/{YAGPT_MODEL_NAME}/latest"

    # 5) Лог-скелет
    connection_log = {
        "engine": "yandexgpt",
        "model": YAGPT_MODEL_NAME,
        "temperature": YAGPT_TEMPERATURE,
        "max_tokens": YAGPT_MAX_TOKENS,
        "frames_selected": materials["summary"]["frames_count"],
        "faces_selected": materials["summary"]["faces_count"],
        "texts_present": materials["summary"]["texts_present"],
        "texts_missing": materials["summary"]["texts_missing"],
        "buyer_experience_found": materials["summary"]["buyer_experience_found"],
        "buyer_experience_chars": materials["summary"]["buyer_experience_chars"],
        "buyer_experience_path": materials["summary"]["buyer_experience_path"],
        "buyer_experience_error": materials["summary"]["buyer_experience_error"],
        "payload_chars": len(context_text),
        "payload_preview": context_text[:payload_preview_len],
        # размеры файлов:
        "file_sizes": {
            "frames": frames_sizes,
            "faces": faces_sizes,
            "texts": text_file_sizes,
            "buyer_experience_bytes": buyer_bytes,
            "totals": {
                "frames_bytes": total_frames_bytes,
                "faces_bytes": total_faces_bytes,
                "texts_bytes": total_text_bytes,
                "all_bytes": total_all_bytes
            }
        },
        "call_path": None,
    }

    # 6) Вызов: SDK → REST fallback
    try:
        try:
            from yandex_cloud_ml_sdk import YCloudML
            sdk = YCloudML(folder_id=yandex_folder_id, auth=yandex_api_key)
            result_obj, call_path = _yagpt_complete_with_sdk(
                sdk, model_uri, messages, YAGPT_TEMPERATURE, YAGPT_MAX_TOKENS
            )
        except Exception as e_sdk:
            call_path = f"sdk_failed:{type(e_sdk).__name__}"
            print(f"⚠️ SDK вызов не удался ({e_sdk}), пробуем REST…")
            result_obj, call_path = _yagpt_complete_with_rest(
                model_uri, messages, YAGPT_TEMPERATURE, YAGPT_MAX_TOKENS
            )

        out_text = _extract_text_from_yagpt_response(result_obj) or "⚠️ Пустой ответ модели."

        header = [
            "Результаты предобработки рекламных материалов",
            f"МОДЕЛЬ: {YAGPT_MODEL_NAME}",
            "РЕЖИМ: текстовый (YandexGPT, сводка всех материалов)",
            f"frames: {materials['summary']['frames_count']} | faces: {materials['summary']['faces_count']}",
            f"Включённые разделы: {', '.join(materials['summary']['texts_present']) or '—'}",
            "----------------------------------------",
        ]
        final_text = "\n".join(header) +"\n"+ out_text

        # Сохраняем
        connection_log["ok"] = True
        connection_log["call_path"] = call_path
        with open(review_path, "w", encoding="utf-8") as f:
            f.write(final_text)
        with open(model_used_path, "w", encoding="utf-8") as f:
            f.write(YAGPT_MODEL_NAME)
        with open(conn_log_path, "w", encoding="utf-8") as f:
            json.dump(connection_log, f, ensure_ascii=False, indent=2)

        print(f"🧾 Лог подключения сохранён: {conn_log_path}")
        print(f"📜 Ответ YandexGPT сохранён: {review_path}")
        print("✅ Анализ завершён.")
        return final_text

    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        print(f"⚠️ Ошибка YandexGPT: {err}")
        connection_log["ok"] = False
        connection_log["error"] = err
        with open(review_path, "w", encoding="utf-8") as f:
            f.write("❌ Не удалось получить ответ от YandexGPT. Подробности — в логе.")
        with open(model_used_path, "w", encoding="utf-8") as f:
            f.write(YAGPT_MODEL_NAME)
        with open(conn_log_path, "w", encoding="utf-8") as f:
            json.dump(connection_log, f, ensure_ascii=False, indent=2)
        return ""
# =============== КОНЕЦ БЛОКА: YANDEXGPT ===============

# =============== НОВЫЙ БЛОК: OpenAI GPT — connectivity + materials analysis ===============
import os
import json
from datetime import datetime
import requests

# ---[ Интеграция: .env, прокси и SDK OpenAI (httpx) ]---
try:
    from dotenv import load_dotenv  # type: ignore
    load_dotenv()
except Exception:
    pass

try:
    import httpx  # type: ignore
except Exception:
    httpx = None

try:
    from openai import OpenAI  # type: ignore
except Exception:
    OpenAI = None

# Используем ту же прокси-конфигурацию, что и выше
PROXY_URL = _OAI_PROXY_URL
# PROXIES уже определён выше и используется повсеместно

# Те же пути, что заданы для Yandex-блока; если их нет — задаём строго.
try:
    PROMPT_GENERAL_FILE
except NameError:
    PROMPT_GENERAL_FILE = r"C:\Users\PTambulatov\Desktop\PT\2.Расчеты\50. ХАКАТОН\12. Prompt\1.Prompt_general.txt"

try:
    BUYER_EXPERIENCE_PATH
except NameError:
    BUYER_EXPERIENCE_PATH = r"C:\Users\PTambulatov\Desktop\PT\2.Расчеты\50. ХАКАТОН\11. Опыт баера\Buyer_Experience.xlsx"

# Ограничители объёма (по аналогии с Yandex-блоком)
try:
    MAX_TEXT_PER_SECTION
except NameError:
    MAX_TEXT_PER_SECTION = 120_000

try:
    MAX_TOTAL_CONTEXT
except NameError:
    MAX_TOTAL_CONTEXT = 280_000

# ---- Параметры OpenAI (модель берём из окружения, задаётся из Flask) ----
OPENAI_MODEL = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
OPENAI_TEMPERATURE = float(os.getenv("OPENAI_TEMPERATURE", "0.2"))
OPENAI_MAX_TOKENS  = int(os.getenv("OPENAI_MAX_TOKENS", "1500"))  # лимит вывода

# 🆕 Поддержка reasoning-моделей
OPENAI_REASONING_EFFORT = os.getenv("OPENAI_REASONING_EFFORT", "medium")  # low|medium|high
REASONING_MODELS = {"o3", "o3-mini", "o3-pro", "o4-mini"}

def _safe_read_text(path: str, limit: int = MAX_TEXT_PER_SECTION) -> str:
    if not path or not os.path.isfile(path):
        return ""
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            with open(path, "r", encoding=enc, errors="strict") as f:
                t = f.read()
            return (t[:limit] + "\n\n…(обрезано)…") if len(t) > limit else t
        except Exception:
            continue
    return ""

def _read_pdf_summary(pdf_path: str, max_pages: int = 3, limit_chars: int = 6000) -> dict:
    """
    Короткое резюме Annotated_Frames.pdf:
    - pages: количество страниц (кадров),
    - text_preview: извлекаемый текст первых страниц (если он там есть),
    - note: подсказка, что это, скорее всего, изображение без текста.
    """
    info = {"exists": False, "pages": 0, "text_preview": "", "note": ""}
    if not pdf_path or not os.path.isfile(pdf_path):
        return info
    info["exists"] = True
    try:
        import pdfplumber  # уже используется выше в файле
        with pdfplumber.open(pdf_path) as pdf:
            info["pages"] = len(pdf.pages)
            chunks = []
            for i, page in enumerate(pdf.pages[:max_pages], start=1):
                txt = page.extract_text() or ""
                if txt.strip():
                    chunks.append(f"[страница {i}]\n{txt.strip()}")
            if chunks:
                preview = ("\n\n".join(chunks))[:limit_chars]
                info["text_preview"] = preview
            else:
                info["note"] = "Текст не извлечён (страницы состоят из изображений с кадрами)."
    except Exception as e:
        info["note"] = f"Ошибка чтения PDF: {e}"
    return info

def _read_buyer_experience_any(path: str, limit: int = MAX_TEXT_PER_SECTION) -> dict:
    """
    Пробуем pandas → openpyxl, как в блоке YandexGPT. Возвращаем текст + метаданные.
    """
    meta = {"path": path, "found": False, "error": None}
    out  = ""
    if not os.path.isfile(path):
        meta["error"] = "file_not_found"
        return {"text": "", "meta": meta}
    # pandas
    try:
        import pandas as pd  # type: ignore
        try:
            sheets = pd.read_excel(path, sheet_name=None, dtype=str)
            buff = []
            for sname, df in sheets.items():
                buff.append(f"Лист: {sname}")
                df = df.iloc[:200, :30].fillna("")
                for row in df.itertuples(index=False, name=None):
                    buff.append("\t".join(str(x) for x in row))
                buff.append("")
            out = "\n".join(buff)
            meta["found"] = True
        except Exception as e_pd:
            meta["error"] = f"pandas_failed: {e_pd}"
    except Exception:
        meta["error"] = meta["error"] or "pandas_not_available"
    # openpyxl fallback
    if not meta["found"]:
        try:
            from openpyxl import load_workbook  # type: ignore
            try:
                wb = load_workbook(path, data_only=True, read_only=True)
                buff = []
                for ws in wb.worksheets:
                    buff.append(f"Лист: {ws.title}")
                    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                        vals = [("" if v is None else str(v)) for v in list(row)[:30]]
                        buff.append("\t".join(vals))
                        if i >= 200:
                            break
                    buff.append("")
                out = "\n".join(buff)
                meta["found"] = True
            except Exception as e_ox:
                meta["error"] = f"openpyxl_failed: {e_ox}"
        except Exception:
            meta["error"] = meta["error"] or "openpyxl_not_available"
    if out and len(out) > limit:
        out = out[:limit] + "\n\n…(обрезано)…"
    return {"text": out, "meta": meta}

def _read_prompt_general_strict_openai() -> str:
    """
    Строго читаем файл промпта (тот же путь, что и для Yandex-блока).
    """
    path = os.path.normpath(PROMPT_GENERAL_FILE)
    for enc in ("utf-8-sig", "utf-8", "cp1251"):
        try:
            with open(path, "r", encoding=enc, errors="strict") as f:
                t = f.read().strip()
            return t
        except Exception:
            continue
    return ""

def _build_openai_context(result_dir: str) -> dict:
    """
    Готовим текстовый контекст строго из требуемых материалов.
    Возвращает {context_text, diag}, где diag — размеры/наличие.
    Теперь вместо Excel встраиваем содержимое Buyer_Experience.json.
    """
    # 1) PDF (Annotated_Frames.pdf) — только краткая сводка (PDF не отправляем)
    annotated_pdf = os.path.join(result_dir, "Annotated_Frames.pdf")
    pdf_info = _read_pdf_summary(annotated_pdf)

    # 2) Текстовые файлы (строго те, что вы просили)
    names = [
        "All_Frames_Text.txt",
        "Documents_Texts.txt",
        "Music.txt",
        "Music_Shazam.txt",
        "Overlay_Report.txt",
        "Speech.txt",
        "Speech_Yandex.txt",
    ]
    texts = {}
    diag_sizes = {}
    for name in names:
        p = os.path.join(result_dir, name)
        t = _safe_read_text(p)
        texts[name] = t
        try:
            diag_sizes[name] = os.path.getsize(p) if os.path.isfile(p) else 0
        except Exception:
            diag_sizes[name] = 0

    # 3) 🆕 Вместо Excel читаем JSON из строгого пути
    buyer_json_path = BUYER_EXPERIENCE_JSON_PATH
    buyer_json_text = _safe_read_text(buyer_json_path)
    buyer_json_bytes = 0
    try:
        buyer_json_bytes = os.path.getsize(buyer_json_path) if os.path.isfile(buyer_json_path) else 0
    except Exception:
        buyer_json_bytes = 0

    # 4) Сборка контекста
    parts = []
    parts.append("=== КРАТКАЯ СВОДКА МАТЕРИАЛОВ (для OpenAI) ===")
    parts.append(f"Annotated_Frames.pdf: {'есть' if pdf_info['exists'] else 'нет'}, страниц: {pdf_info.get('pages', 0)}.")
    if pdf_info.get("text_preview"):
        parts.append("Короткий извлекаемый текст из первых страниц PDF:")
        parts.append(pdf_info["text_preview"])
    elif pdf_info.get("note"):
        parts.append(pdf_info["note"])
    parts.append("")

    for nm in names:
        if texts[nm]:
            parts.append(f"=== {nm} ===")
            parts.append(texts[nm])
            parts.append("")

    # 🆕 Вкладываем именно JSON (а не текстовую «распечатку» Excel)
    parts.append("=== Buyer_Experience.json (машиночитаемый чек-лист) ===")
    if buyer_json_text:
        parts.append(buyer_json_text)
    else:
        parts.append(f"[Не найден {buyer_json_path} или пустой — OpenAI-блок продолжит без чек-листа]")

    context = "\n".join(parts)
    if len(context) > MAX_TOTAL_CONTEXT:
        context = context[:MAX_TOTAL_CONTEXT] + "\n\n…(контекст укорочен)…"

    diag = {
        "pdf": {"path": annotated_pdf, "exists": pdf_info["exists"], "pages": pdf_info.get("pages", 0)},
        "text_file_bytes": diag_sizes,
        "buyer_experience_json": {
            "path": buyer_json_path,
            "bytes": buyer_json_bytes,
            "found": bool(buyer_json_bytes)
        },
        "payload_chars": len(context)
    }
    return {"context_text": context, "diag": diag}

def _openai_sdk_models_probe():
    """
    Возвращает dict: {ok, total_models, models, error}
    """
    api_key = os.getenv("OPENAI_API_KEY", "")
    if not api_key or OpenAI is None:
        return {"ok": False, "total_models": 0, "models": [], "error": "SDK недоступен или ключ не задан"}
    try:
        if PROXY_URL and httpx is not None:
            client = OpenAI(http_client=httpx.Client(proxy=PROXY_URL), api_key=api_key)
        else:
            client = OpenAI(api_key=api_key)
        resp = client.models.list()
        items = getattr(resp, "data", []) or []
        model_ids = []
        for m in items:
            mid = getattr(m, "id", None)
            if mid is None and isinstance(m, dict):
                mid = m.get("id")
            if mid is not None:
                model_ids.append(str(mid))
        return {"ok": True, "total_models": len(model_ids), "models": model_ids, "error": None}
    except Exception as e:
        return {"ok": False, "total_models": 0, "models": [], "error": str(e)}


# === OpenAI — PING через Responses API =======================================
def _call_openai_ping():
    """
    Мини-запрос к OpenAI Responses API, чтобы проверить доступность:
    - отправляем system+user с просьбой ответить "pong"
    - возвращаем краткую структуру с ok/http_status/фрагментом ответа
    """
    import os, json, requests

    api_key = os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        return {"ok": False, "http_status": None, "error": "OPENAI_API_KEY is missing"}

    model = os.getenv("OPENAI_MODEL", "gpt-4o-mini")
    url = "https://api.openai.com/v1/responses"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}

    # Формируем input в формате Responses API: список ролей с контент-блоками
    input_messages = [
        {"role": "system", "content": [{"type": "input_text", "text": "You are a helpful assistant."}]},
        {"role": "user",   "content": [{"type": "input_text", "text": "Ответь одним словом: pong."}]}
    ]

    payload = {
        "model": model,
        "input": input_messages,
        "temperature": 0.0,
    }

    # Для reasoning-моделей добавим effort — в Responses это то же поле
    if model in REASONING_MODELS:
        payload["reasoning"] = {"effort": os.getenv("OPENAI_REASONING_EFFORT", OPENAI_REASONING_EFFORT)}
        payload["max_output_tokens"] = 32
    else:
        payload["max_output_tokens"] = 32

    try:
        r = requests.post(url, headers=headers, json=payload, timeout=15, proxies=PROXIES)
        status = r.status_code
        if status == 200:
            j = r.json()
            # Responses API: берём output_text; если нет — выжимаем текст из output/choices и т.п.
            text = _extract_text_from_openai_responses(j) or json.dumps(j, ensure_ascii=False)[:300]
            return {"ok": True, "http_status": status, "model": model, "reply_excerpt": (text or "").strip()[:200]}
        else:
            try:
                err = r.json()
            except Exception:
                err = {"raw": r.text[:500]}
            return {"ok": False, "http_status": status, "model": model, "error": err}
    except Exception as e:
        return {"ok": False, "http_status": None, "model": model, "error": str(e)}


def _safe_get_json(url, params=None, headers=None, timeout=8):
    try:
        r = requests.get(url, params=params, headers=headers, timeout=timeout, proxies=PROXIES)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        return {"error": str(e), "url": url}

def _detect_ip_and_country():
    data = _safe_get_json("https://ipapi.co/json")
    if isinstance(data, dict) and ("ip" in data or "country_name" in data):
        return {
            "ip": data.get("ip"),
            "country_name": data.get("country_name"),
            "country_code": data.get("country"),
            "source": "ipapi.co",
            "raw": data
        }
    ipify = _safe_get_json("https://api.ipify.org?format=json")
    if isinstance(ipify, dict) and "ip" in ipify:
        ip = ipify["ip"]
        data2 = _safe_get_json(f"https://ipinfo.io/{ip}/json")
        return {
            "ip": ip,
            "country_name": data2.get("country") if isinstance(data2, dict) else None,
            "country_code": data2.get("country") if isinstance(data2, dict) else None,
            "source": "ipify+ipinfo",
            "raw": {"ipify": ipify, "ipinfo": data2}
        }
    return {"ip": None, "country_name": None, "country_code": None, "source": "unavailable", "raw": {}}



# ===== YANDEX GPT — CONNECTIVITY TEST (IP, страна, пинг) =====================
import os, json, requests
from datetime import datetime

def _ygpt_safe_get_json(url, params=None, headers=None, timeout=8, proxies=None):
    try:
        r = requests.get(url, params=params, headers=headers, timeout=timeout, proxies=proxies)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        return {"error": str(e), "url": url}

def _ygpt_detect_ip_and_country(proxies=None):
    data = _ygpt_safe_get_json("https://ipapi.co/json", proxies=proxies)
    if isinstance(data, dict) and ("ip" in data or "country_name" in data):
        return {
            "ip": data.get("ip"),
            "country_name": data.get("country_name"),
            "country_code": data.get("country"),
            "source": "ipapi.co",
            "raw": data
        }
    ipify = _ygpt_safe_get_json("https://api.ipify.org?format=json", proxies=proxies)
    if isinstance(ipify, dict) and "ip" in ipify:
        ip = ipify["ip"]
        data2 = _ygpt_safe_get_json(f"https://ipinfo.io/{ip}/json", proxies=proxies)
        return {
            "ip": ip,
            "country_name": data2.get("country") if isinstance(data2, dict) else None,
            "country_code": data2.get("country") if isinstance(data2, dict) else None,
            "source": "ipify+ipinfo",
            "raw": {"ipify": ipify, "ipinfo": data2}
        }
    return {"ip": None, "country_name": None, "country_code": None, "source": "unavailable", "raw": {}}

def _yandex_gpt_ping(api_key: str, folder_id: str, proxies=None):
    """
    Мини-запрос к YandexGPT (non-stream) для «пинга».
    Док: https://cloud.yandex.ru/docs/foundation-models/
    """
    url = "https://llm.api.cloud.yandex.net/foundationModels/v1/completion"
    # Модель: lite быстрее/дешевле и отлично подходит для пинга.
    model = f"gpt://{folder_id}/yandexgpt-lite/latest"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Api-Key {api_key}",
    }
    payload = {
        "model": model,
        "completionOptions": {"stream": False, "temperature": 0.1, "maxTokens": 32},
        "messages": [
            {"role": "system", "text": "Ты — минимальный сетевой зонд. Отвечай одним словом на русском."},
            {"role": "user",   "text": "Ответь словом: Ping."}
        ]
    }
    r = requests.post(url, headers=headers, json=payload, timeout=20, proxies=proxies)
    status = r.status_code
    try:
        data = r.json()
    except Exception:
        data = {"raw": r.text[:500]}
    if status == 200:
        try:
            reply = data["result"]["alternatives"][0]["message"]["text"].strip()
        except Exception:
            reply = json.dumps(data, ensure_ascii=False)[:200]
        return {"ok": True, "http_status": status, "reply": reply, "raw": data}
    else:
        return {"ok": False, "http_status": status, "error": data}

def run_yandex_connectivity_test(result_dir: str):
    """
    Пишет 2 файла:
      - YandexGPT_Connectivity_Test.txt   (короткий читаемый вывод)
      - YandexGPT_Connectivity_Log.json   (детальный лог)
    Требуются переменные окружения: YANDEX_API_KEY, YANDEX_FOLDER_ID.
    Необязательная: OPENAI_PROXY_URL (исп. как системный proxy и для этого пинга).
    """
    out_txt = os.path.join(result_dir, "YandexGPT_Connectivity_Test.txt")
    out_log = os.path.join(result_dir, "YandexGPT_Connectivity_Log.json")

    api_key = os.getenv("YANDEX_API_KEY", "")
    folder_id = os.getenv("YANDEX_FOLDER_ID", "")
    proxy_url = os.getenv("OPENAI_PROXY_URL") or ""  # используем, если уже настроен общий прокси
    proxies = {"http": proxy_url, "https": proxy_url} if proxy_url else None

    ts_utc = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    geo = _ygpt_detect_ip_and_country(proxies=proxies)

    short_lines = []
    short_lines.append(f"Публичный IP: {geo.get('ip') or '—'}")
    cn = geo.get("country_name"); cc = geo.get("country_code")
    if cn and cc and cn != cc:
        short_lines.append(f"Страна: {cn} ({cc})")
    elif cn or cc:
        short_lines.append(f"Страна: {cn or cc}")
    else:
        short_lines.append("Страна: —")

    if not api_key or not folder_id:
        short_lines.append("YandexGPT API: ОШИБКА")
        short_lines.append("Подробности: отсутствуют YANDEX_API_KEY или YANDEX_FOLDER_ID")
        text_out = "\n".join(short_lines)
        try:
            with open(out_txt, "w", encoding="utf-8") as f:
                f.write(text_out + "\n")
            with open(out_log, "w", encoding="utf-8") as f:
                json.dump({"timestamp_utc": ts_utc, "geo": geo,
                           "error": "missing_env", "required": ["YANDEX_API_KEY", "YANDEX_FOLDER_ID"]},
                          f, ensure_ascii=False, indent=2)
        except Exception:
            pass
        print(f"⚠️ YandexGPT connectivity: {out_txt}")
        return text_out

    # Пинг YandexGPT
    ping = _yandex_gpt_ping(api_key=api_key, folder_id=folder_id, proxies=proxies)
    if ping.get("ok"):
        short_lines.append(f"YandexGPT API: OK , HTTP: {ping.get('http_status')}")
        short_lines.append(f"Ответ модели: {ping.get('reply') or '—'}")
    else:
        short_lines.append("YandexGPT API: ОШИБКА" + (f" , HTTP: {ping.get('http_status')}" if ping.get("http_status") else ""))
        short_lines.append(f"Подробности: {str(ping.get('error'))[:600]}")

    text_out = "\n".join(short_lines).strip()

    # Записи файлов
    try:
        with open(out_txt, "w", encoding="utf-8") as f:
            f.write(text_out + "\n")
    except Exception as e:
        text_out += f"\n[Ошибка записи TXT: {e}]"

    log = {
        "timestamp_utc": ts_utc,
        "env": {"YANDEX_FOLDER_ID_present": bool(folder_id), "proxy_present": bool(proxy_url)},
        "geo": geo,
        "yandex_ping": ping
    }
    try:
        with open(out_log, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    print(f"🔌 Результат YandexGPT connectivity: {out_txt}")
    return text_out
# =======================================================================



def run_openai_connectivity_test(result_dir: str):
    """
    Определение IP/страны, пинг, список моделей,
    запись краткого вывода в OpenAI_Connectivity_Test.txt, детальный лог — в JSON.
    """
    out_txt = os.path.join(result_dir, "OpenAI_Connectivity_Test.txt")
    out_log = os.path.join(result_dir, "OpenAI_Connectivity_Log.json")

    ts_utc = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC")
    geo = _detect_ip_and_country()
    ping = _call_openai_ping()
    sdk_probe = _openai_sdk_models_probe()

    summary_lines = []
    summary_lines.append(f"Публичный IP: {geo.get('ip') or '—'}")
    cn = geo.get("country_name")
    cc = geo.get("country_code")
    if cn and cc and cn != cc:
        summary_lines.append(f"Страна: {cn} ({cc})")
    elif cn or cc:
        summary_lines.append(f"Страна: {cn or cc}")
    else:
        summary_lines.append("Страна: —")

    if ping.get("ok"):
        summary_lines.append(f"OpenAI API: OK , HTTP: {ping.get('http_status')}")
        excerpt = (ping.get("reply_excerpt") or "").replace("\n", " ").strip()
        summary_lines.append(f"Ответ модели: {excerpt or '—'}")
        if sdk_probe.get("ok"):
            summary_lines.append(f"Доступных моделей: {sdk_probe.get('total_models', 0)}")
        else:
            summary_lines.append("Доступных моделей: —")
    else:
        http = ping.get("http_status")
        line = "OpenAI API: ОШИБКА"
        if http is not None:
            line += f" , HTTP: {http}"
        summary_lines.append(line)
        summary_lines.append(f"Подробности: {str(ping.get('error'))[:1000]}")

    text_out = "\n".join(summary_lines).strip()

    try:
        with open(out_txt, "w", encoding="utf-8") as f:
            f.write(text_out + "\n")
    except Exception as e:
        text_out += f"\n[Ошибка записи TXT: {e}]"

    ip_s = geo.get("ip") or "—"
    country_disp = (cn and cc and f"{cn} ({cc})") or (cn or cc or "—")
    models_list = sdk_probe.get("models") or []
    models_count = sdk_probe.get("total_models", len(models_list) if models_list else 0)
    connection_details_line = f"ip={ip_s}; country={country_disp}; models_count={models_count}; models=[{', '.join(models_list) if models_list else '—'}]"

    log = {
        "timestamp_utc": ts_utc,
        "env": {
            "OPENAI_MODEL": OPENAI_MODEL,
            "OPENAI_PROXY_URL_present": bool(PROXY_URL),
        },
        "geo": geo,
        "openai_ping": ping,
        "openai_sdk_probe": sdk_probe,
        "connection_details_line": connection_details_line
    }
    try:
        with open(out_log, "w", encoding="utf-8") as f:
            json.dump(log, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

    print(f"🔌 Результат OpenAI connectivity: {out_txt}")
    return text_out



def _openai_responses_complete(messages: list, model: str = OPENAI_MODEL,
                               temperature: float = OPENAI_TEMPERATURE,
                               max_tokens: int = OPENAI_MAX_TOKENS) -> dict:
    """
    Универсальный вызов OpenAI Responses API через requests.
    При 403/404/422 — автоматический fallback на Chat Completions.
    Всегда идём через один и тот же прокси (PROXIES).
    """
    import os, requests

    api_key = os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY отсутствует в окружении")

    url = "https://api.openai.com/v1/responses"
    headers = {"Authorization": f"Bearer {api_key}", "Content-Type": "application/json"}

    input_messages = _to_responses_input_from_chat_messages(messages)
    payload = {"model": model, "input": input_messages, "temperature": float(temperature)}

    max_out = max(int(max_tokens), 16)
    if model in REASONING_MODELS:
        payload["reasoning"] = {"effort": os.getenv("OPENAI_REASONING_EFFORT", OPENAI_REASONING_EFFORT)}
    payload["max_output_tokens"] = int(max_out)

    r = requests.post(url, headers=headers, json=payload, timeout=90, proxies=PROXIES)

    # Явный fallback для блокировок/несовместимостей
    if r.status_code in (403, 404, 422):
        return _openai_chat_completions_fallback(messages, model, temperature, max_out, headers)

    r.raise_for_status()
    return r.json()



def _openai_chat_completions_fallback(messages: list, model: str, temperature: float,
                                      max_tokens: int, headers: dict) -> dict:
    """
    Если /v1/responses недоступен/запрещён (403/404/422), идём на /v1/chat/completions
    с теми же сообщениями и через тот же прокси.
    Возвращаем объект с полем output_text для унифицированного парсинга.
    """
    import requests
    payload = {
        "model": model,
        "messages": [{"role": m.get("role", "user"), "content": m.get("content", "")} for m in messages],
        "temperature": float(temperature),
        "max_tokens": int(max_tokens)
    }
    r = requests.post(
        "https://api.openai.com/v1/chat/completions",
        headers=headers, json=payload, timeout=90, proxies=PROXIES
    )
    r.raise_for_status()
    data = r.json()
    try:
        text = (data["choices"][0]["message"]["content"] or "").strip()
    except Exception:
        text = json.dumps(data, ensure_ascii=False)[:2000]
    return {"output_text": text, "raw": data}





def _to_responses_input_from_chat_messages(messages: list) -> list:
    """
    Преобразует старые chat-сообщения в формат Responses API:
    [{"role": "...", "content": [{"type": "input_text", "text": "..."}]}, ...]
    """
    out = []
    for m in messages or []:
        role = m.get("role") or "user"
        content = m.get("content") or ""
        if isinstance(content, str):
            parts = [{"type": "input_text", "text": content}]
        elif isinstance(content, list):
            # если уже пришёл массив блоков — аккуратно нормализуем типы text -> input_text
            parts = []
            for c in content:
                if isinstance(c, dict):
                    ctype = c.get("type")
                    if ctype == "text":  # ← старый вариант
                        parts.append({"type": "input_text", "text": c.get("text", "")})
                    else:
                        parts.append(c)
                else:
                    parts.append({"type": "input_text", "text": str(c)})
        else:
            parts = [{"type": "input_text", "text": str(content)}]
        out.append({"role": role, "content": parts})
    return out


def _extract_text_from_openai_responses(resp: dict) -> str:
    """
    Аккуратно извлекает текст из ответа Responses API.
    Приоритет: output_text -> response.output[...].content[...].text -> choices/message/content -> сырой json
    """
    try:
        if not isinstance(resp, dict):
            return str(resp)

        # 1) Прямо в Responses обычно есть "output_text"
        t = resp.get("output_text")
        if isinstance(t, str) and t.strip():
            return t.strip()

        # 2) Универсальный путь через "output"
        out = resp.get("output")
        if isinstance(out, list) and out:
            content = out[0].get("content")
            if isinstance(content, list):
                for c in content:
                    if isinstance(c, dict) and c.get("type") in ("output_text", "text"):
                        txt = c.get("text")
                        if isinstance(txt, str) and txt.strip():
                            return txt.strip()

        # 3) На всякий случай поддержим ответы в стиле choices/message/content (если шлюз изменится)
        if "choices" in resp:
            choices = resp.get("choices") or []
            if choices:
                msg = choices[0].get("message") or {}
                ct = msg.get("content")
                if isinstance(ct, str) and ct.strip():
                    return ct.strip()

        # 4) Ничего не нашли — вернём компактный json
        return json.dumps(resp, ensure_ascii=False)[:2000]
    except Exception:
        return json.dumps(resp, ensure_ascii=False)[:2000]




# === OpenAI tokens status (usage) ============================================
from datetime import date, datetime

def _oai_fetch_usage(start_dt: date, end_dt: date):
    """
    Пытается получить usage с https://api.openai.com/v1/usage
    Возвращает dict (JSON) или dict с ключом 'error' при неудаче.
    """
    api_key = os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        return {"error": "OPENAI_API_KEY is missing"}
    try:
        url = "https://api.openai.com/v1/usage"
        headers = {"Authorization": f"Bearer {api_key}"}
        params = {
            "start_date": start_dt.strftime("%Y-%m-%d"),
            "end_date":   end_dt.strftime("%Y-%m-%d"),
        }
        r = requests.get(url, headers=headers, params=params, timeout=30, proxies=PROXIES)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        return {"error": f"{type(e).__name__}: {e}"}

def _oai_try_extract_tokens_total(usage_json: dict) -> int | None:
    """
    Универсальная попытка посчитать суммарные токены из разных возможных форматов ответа.
    Если формат неизвестен — вернём None (UI покажет «нет данных»).
    """
    try:
        if not isinstance(usage_json, dict):
            return None
        # Популярные поля, встречающиеся в разных сводках
        total = 0
        hit = False
        # Вариант 1: список дней
        data = usage_json.get("data")
        if isinstance(data, list) and data:
            for item in data:
                for k in ("n_context_tokens_total", "n_output_tokens_total",
                          "n_generated_tokens_total", "total_tokens"):
                    v = item.get(k)
                    if isinstance(v, (int, float)):
                        total += int(v)
                        hit = True
            if hit:
                return total
        # Вариант 2: агрегаты
        agg = usage_json.get("aggregated") or usage_json.get("total_usage") or {}
        for k in ("n_context_tokens_total", "n_output_tokens_total",
                  "n_generated_tokens_total", "total_tokens"):
            v = agg.get(k)
            if isinstance(v, (int, float)):
                total += int(v)
                hit = True
        return total if hit else None
    except Exception:
        return None

def write_openai_tokens_status(result_dir: str):
    """
    Пишет файл Clipchecker_materials/OpenAI_Tokens_Status.json
    Структура:
      {
        "timestamp_utc": "...",
        "limits": {"month_tokens": int|None, "day_tokens": int|None},
        "used":   {"month_tokens": int|None, "day_tokens": int|None},
        "period": {"month_start": "YYYY-MM-DD", "today": "YYYY-MM-DD"},
        "source": "openai_usage_api"|"fallback",
        "errors": {"month": "...", "day": "..."}    # при сбоях
      }
    Значения лимитов берутся из окружения: OPENAI_LIMIT_TOKENS_MONTH / OPENAI_LIMIT_TOKENS_DAY (необяз.)
    """
    out_json = os.path.join(result_dir, "OpenAI_Tokens_Status.json")

    # Периоды
    today = date.today()
    month_start = today.replace(day=1)

    # Лимиты (если заданы пользователем)
    lim_month = os.getenv("OPENAI_LIMIT_TOKENS_MONTH")
    lim_day   = os.getenv("OPENAI_LIMIT_TOKENS_DAY")
    try:
        lim_month_int = int(lim_month) if lim_month else None
    except Exception:
        lim_month_int = None
    try:
        lim_day_int = int(lim_day) if lim_day else None
    except Exception:
        lim_day_int = None

    # Запрашиваем usage
    u_month = _oai_fetch_usage(month_start, today)
    u_day   = _oai_fetch_usage(today, today)

    used_month = None if isinstance(u_month, dict) and u_month.get("error") else _oai_try_extract_tokens_total(u_month)
    used_day   = None if isinstance(u_day, dict)   and u_day.get("error")   else _oai_try_extract_tokens_total(u_day)

    payload = {
        "timestamp_utc": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
        "limits": {"month_tokens": lim_month_int, "day_tokens": lim_day_int},
        "used":   {"month_tokens": used_month, "day_tokens": used_day},
        "period": {"month_start": month_start.isoformat(), "today": today.isoformat()},
        "source": "openai_usage_api",
        "errors": {
            "month": u_month.get("error") if isinstance(u_month, dict) and u_month.get("error") else None,
            "day":   u_day.get("error")   if isinstance(u_day, dict)   and u_day.get("error")   else None,
        }
    }
    # Если оба запроса провалились — всё равно создадим файл
    if payload["errors"]["month"] and payload["errors"]["day"]:
        payload["source"] = "fallback"

    try:
        with open(out_json, "w", encoding="utf-8") as f:
            json.dump(payload, f, ensure_ascii=False, indent=2)
        print(f"🧾 OpenAI_Tokens_Status.json создан: {out_json}")
    except Exception as e:
        # Как минимум — создадим очень простой файл с ошибкой
        try:
            with open(out_json, "w", encoding="utf-8") as f:
                json.dump({"error": f"failed_to_write: {e}"}, f, ensure_ascii=False, indent=2)
            print(f"🧾 OpenAI_Tokens_Status.json создан с ошибкой записи: {out_json}")
        except Exception:
            print("❌ Не удалось записать OpenAI_Tokens_Status.json")



def _extract_text_from_openai_response(resp: dict) -> str:
    try:
        return (resp["choices"][0]["message"]["content"] or "").strip()
    except Exception:
        return json.dumps(resp, ensure_ascii=False)




# 🔄 NEW helper: кодирует локальный файл-кадр в data: URL для Responses API
def _encode_image_as_data_url(path: str) -> str:
    import base64, mimetypes
    mime = (mimetypes.guess_type(path)[0]) or "image/jpeg"
    with open(path, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("ascii")
    return f"data:{mime};base64,{b64}"





def run_openai_materials_review(result_dir: str):

    """
    Анализ материалов в OpenAI (ChatGPT) БЕЗ PDF:
    - НЕ отправляем Annotated_Frames.pdf / Faces_*.pdf.
    - Добавляем кадры из .../Clipchecker_materials/frames как input_image (data:URL).
    - Логируем статус изображений и явный запрет на PDF.
    - Сохраняем ответ в Comprehensive_Legal_Review_OpenAI.txt,
      диагностику — в OpenAI_Materials_Log.json,
      историю — в OpenAI_Chat_History.json.
    """
    import os, json, base64, mimetypes, datetime

    # --- локальные хелперы, чтобы не зависеть от внешних _write_* ---
    def _wt(path: str, text: str):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            f.write(text)

    def _wj(path: str, data: dict):
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)

    def _append_log(message: str):
        try:
            log_path = os.path.join(result_dir, "clipchecker.log")
            ts = datetime.datetime.now().strftime("%H:%M:%S")
            with open(log_path, "a", encoding="utf-8") as f:
                f.write(f"[{ts}] {message}\n")
        except Exception:
            pass
    # ----------------------------------------------------------------

    out_txt = os.path.join(result_dir, "Comprehensive_Legal_Review_OpenAI.txt")
    out_log = os.path.join(result_dir, "OpenAI_Materials_Log.json")
    chat_hist_path = os.path.join(result_dir, "OpenAI_Chat_History.json")

    # 1) Сбор текстового контекста (как раньше)
    bundle = _build_openai_context(result_dir)
    context_text = bundle.get("context_text", "")
    diag = bundle.get("diag", {}) or {}

    # 🆕 ЛОГ: подтверждаем, что JSON попал в контекст OpenAI
    _append_log(
        f"🧾 В OpenAI контекст добавлён Buyer_Experience.json "
        f"({diag.get('buyer_experience_json', {}).get('bytes', 0)} байт)"
    )    

    # 2) Загружаем общий промпт
    prompt_general = _read_prompt_general_strict_openai()
    if not prompt_general:
        msg = ("❌ Не удалось прочитать текст промпта из 1.Prompt_general.txt по строгому пути.\n"
               f"Ожидался файл: {PROMPT_GENERAL_FILE}")
        _wt(out_txt, msg)
        _wj(out_log, {"error": "prompt_general_not_found", "path": PROMPT_GENERAL_FILE, "diag": diag})
        print(msg)
        _append_log("⛔ OpenAI: промпт не найден, анализ прерван.")
        return msg

    # 3) ЖЁСТКО отключаем любые PDF и логируем это
    skipped_pdfs = []
    for name in ("Annotated_Frames.pdf", "Faces_Thumbnails.pdf", "Faces_Strip.pdf"):
        p = os.path.join(result_dir, name)
        if os.path.exists(p):
            skipped_pdfs.append(name)
    diag["pdf_policy"] = {
        "send_pdf": False,
        "skipped_pdfs_present": skipped_pdfs
    }
    if skipped_pdfs:
        _append_log(f"🛑 PDF отправка отключена. Пропущены: {', '.join(skipped_pdfs)}")
    else:
        _append_log("🛑 PDF отправка отключена. В каталоге нет ожидаемых PDF.")

    # 4) Выбираем кадры из frames/
    frames_dir = os.path.join(result_dir, "frames")
    # параметры выборки можно задать через ENV
    try:
        frame_step = int(os.environ.get("OPENAI_CHAT_FRAMES_STEP", "3"))
    except Exception:
        frame_step = 3
    try:
        max_frames = int(os.environ.get("OPENAI_CHAT_FRAMES_MAX", "60"))
    except Exception:
        max_frames = 60
    frame_step = max(1, min(frame_step, 30))
    max_frames = max(1, min(max_frames, 200))

    all_files = []
    selected_paths = []
    included_basenames = []
    errors_images = []
    total_found = 0

    try:
        if os.path.isdir(frames_dir):
            all_files = sorted([
                f for f in os.listdir(frames_dir)
                if f.lower().endswith((".jpg", ".jpeg", ".png", ".webp"))
            ])
            total_found = len(all_files)
            for i in range(0, total_found, frame_step):
                selected_paths.append(os.path.join(frames_dir, all_files[i]))
            if len(selected_paths) > max_frames:
                selected_paths = selected_paths[:max_frames]
        else:
            errors_images.append({"stage": "check_dir", "error": f"frames dir not found: {frames_dir}"})
    except Exception as e:
        errors_images.append({"stage": "listdir", "error": f"{type(e).__name__}: {e}"})

    # 5) Кодируем выбранные кадры как data URL
    def _to_data_url(p: str) -> str:
        mime, _ = mimetypes.guess_type(p)
        if not mime:
            mime = "image/jpeg"
        with open(p, "rb") as fh:
            b64 = base64.b64encode(fh.read()).decode("ascii")
        return f"data:{mime};base64,{b64}"

    image_blocks = []
    for p in selected_paths:
        try:
            data_url = _to_data_url(p)
            image_blocks.append({"type": "input_image", "image_url": data_url})
            included_basenames.append(os.path.basename(p))
        except Exception as e:
            errors_images.append({"file": p, "error": f"{type(e).__name__}: {e}"})

    # Логируем статус изображений
    if included_basenames:
        preview_list = ", ".join(included_basenames[:15])
        tail = "" if len(included_basenames) <= 15 else f" … +{len(included_basenames)-15}"
    else:
        preview_list, tail = "(нет)", ""
    _append_log(
        f"📸 OpenAI кадры: найдено {total_found}, отобрано {len(included_basenames)} (шаг={frame_step}, лимит={max_frames}): {preview_list}{tail}"
    )
    if errors_images:
        _append_log(f"⛔ Ошибки подготовки изображений: {len(errors_images)} (первые 3 в JSON-логе)")
    # пишем в диагностику детальный статус
    diag["images"] = {
        "frames_dir": frames_dir,
        "found_total": total_found,
        "selected_count": len(included_basenames),
        "selected": included_basenames,
        "errors": errors_images,
        "frame_step": frame_step,
        "max_frames": max_frames
    }

    # 6) Сообщения для Responses API: system + user(text + images)
    SYSTEM_TEXT = (
        "Ты делаешь предварительный скрининг рекламных материалов для ТВ. Отвечай по-русски, по чек-листу, "
        "без дисклеймеров и повторов. Запрещено использовать формулировки вроде «Недостаточно данных», "
        "«требуется полная информация» или просьбы прислать ещё документы. "
        "Если каких-то материалов не хватает — всё равно делай выводы по имеющимся данным и "
        "в конце добавь отдельный раздел «Чего не хватает» (списком). "
        "Форматируй вывод компактно, с подзаголовками."
    )

    user_blocks = [
        {"type": "input_text", "text": f"{prompt_general}\n\n---\nМАТЕРИАЛЫ ДЛЯ АНАЛИЗА:\n{context_text}"},
    ]
    # добавим понятный заголовок перед кадрами (если есть)
    if included_basenames:
        user_blocks.append({"type": "input_text", "text": f"Ниже приложены {len(included_basenames)} кадров из папки frames (без PDF):"})
        user_blocks += image_blocks

    messages = [
        {"role": "system", "content": [{"type": "input_text", "text": SYSTEM_TEXT}]},
        {"role": "user",   "content": user_blocks}
    ]

    # 7) Вызов OpenAI через существующий универсальный хелпер (идёт через тот же прокси)
    try:
        resp = _openai_responses_complete(
            messages, model=OPENAI_MODEL,
            temperature=OPENAI_TEMPERATURE, max_tokens=OPENAI_MAX_TOKENS
        )
        out_text = _extract_text_from_openai_responses(resp) or "⚠️ Пустой ответ модели."

        header = [
            "Результаты предобработки рекламных материалов",
            f"МОДЕЛЬ: {OPENAI_MODEL}",
            "РЕЖИМ: мультимодальный (текст + кадры из frames/, БЕЗ PDF)",
            f"картинок в запросе: {len(included_basenames)}",
            "----------------------------------------",
        ]
        final_text = "\n".join(header) + "\n" + out_text
        _wt(out_txt, final_text)

        # Диагностика + старт истории чата
        diag["ok"] = True
        diag["model"] = OPENAI_MODEL
        diag.setdefault("pdf_policy", {}).update({"actually_sent_pdf": False})
        _wj(out_log, diag)

        chat_history = {
            "model": OPENAI_MODEL,
            "messages": messages + [
                {"role": "assistant", "content": [{"type": "input_text", "text": out_text}]}
            ]
        }
        _wj(chat_hist_path, chat_history)

        print(f"💬 Стартовая история чата сохранена: {chat_hist_path}")
        print(f"🧾 OpenAI_Materials_Log.json создан: {out_log}")
        print(f"📜 Ответ OpenAI сохранён: {out_txt}")
        _append_log("✅ OpenAI: анализ выполнен. PDF не отправлялись; учтены кадры из frames.")
        return final_text

    except Exception as e:
        err = f"{type(e).__name__}: {e}"
        diag["ok"] = False
        diag["error"] = err
        diag.setdefault("pdf_policy", {}).update({"actually_sent_pdf": False})
        _wt(out_txt, "❌ Не удалось получить ответ от OpenAI (ChatGPT). Подробности — в OpenAI_Materials_Log.json.")
        _wj(out_log, diag)
        _append_log(f"⛔ OpenAI ошибка: {err}")
        print(f"⚠️ Ошибка OpenAI: {err}")
        return ""





# =============== КОНЕЦ НОВОГО БЛОКА: OpenAI GPT — connectivity + materials analysis ==============


# ====== Safe wrappers: выполняют шаг и при ошибке пишут файл с расшифровкой ======
def run_yagpt_safe(result_dir):
    """
    Пытается выполнить YandexGPT. При ошибке записывает код+расшифровку
    в Comprehensive_Legal_Review.txt (его же читает Flask).
    """
    out_path = os.path.join(result_dir, "Comprehensive_Legal_Review.txt")
    try:
        # ваш основной шаг:
        run_yagpt(result_dir)
    except Exception as e:
        code, msg = _extract_code_and_message_from_response_like(getattr(e, "response", None))
        if not code and not msg:
            code, msg = _extract_code_and_message_from_response_like(e)
        explanation = explain_yandex_error(code or "")
        _write_error_file(out_path, "YandexGPT", code, msg, explanation,
                          extra={"traceback": traceback.format_exc()})

def run_openai_connectivity_test_safe(result_dir):
    """
    Вместо падения пишет человекочитаемый отчёт в OpenAI_Connectivity_Test.txt
    (его же показывает Flask как «пинг»).
    """
    out_path = os.path.join(result_dir, "OpenAI_Connectivity_Test.txt")
    try:
        run_openai_connectivity_test(result_dir)
    except Exception as e:
        code, msg = _extract_code_and_message_from_response_like(getattr(e, "response", None))
        if not code and not msg:
            code, msg = _extract_code_and_message_from_response_like(e)
        explanation = explain_openai_error(code or "")
        _write_error_file(out_path, "OpenAI", code, msg, explanation,
                          extra={"traceback": traceback.format_exc()})
def run_openai_materials_review_safe(result_dir):
    """
    Анализ материалов в OpenAI (ChatGPT). Варианты:
    - Если внутренний код сам записал «заглушку» в Comprehensive_Legal_Review_OpenAI.txt
      и не бросил исключение, мы всё равно подменим её на содержимое лога.
    - Если бросил исключение, попробуем достать «человеческую» строку из лога,
      иначе запишем код/расшифровку.
    """
    out_path = os.path.join(result_dir, "Comprehensive_Legal_Review_OpenAI.txt")
    log_json_path = os.path.join(result_dir, "OpenAI_Materials_Log.json")

    try:
        run_openai_materials_review(result_dir)
        # Даже если исключения не было — возможно, внутренняя функция записала заглушку.
        _maybe_replace_openai_placeholder_with_log(out_path, result_dir)
        return
    except Exception as e:
        # Пытаемся вытащить «живую» строку из лога:
        human_line = _extract_human_error_from_openai_log(log_json_path)
        if not human_line:
            # Если не получилось — применяем общий механизм с кодом/расшифровкой
            code, msg = _extract_code_and_message_from_response_like(getattr(e, "response", None))
            if not code and not msg:
                code, msg = _extract_code_and_message_from_response_like(e)
            explanation = explain_openai_error(code or "")
            _write_error_file(out_path, "OpenAI", code, msg, explanation,
                              extra={"traceback": traceback.format_exc()})
            return

        # Нашли понятное сообщение — пишем его прямо в результирующий файл
        _write_text(out_path, f"Ошибка OpenAI: {human_line}")






if __name__ == "__main__":
    # 1) Финальный шаг «левой» колонки: отчёт по набивкам
    ocr_log_path = os.path.join(result_dir, "OCR_Log.json")
    frames_dir = os.path.join(result_dir, "frames")
    output_path = os.path.join(result_dir, "Overlay_Report.txt")
    write_overlay_report(ocr_log_path, frames_dir, output_path)
    print(f"📝 Отчёт по набивке сохранён: {output_path}")

    # 2) Последовательность шагов по вашему ТЗ:

    # 2.1) YandexGPT временно отключён. Включение: ENABLE_YANDEXGPT=1.
    if os.getenv("ENABLE_YANDEXGPT", "").strip().lower() in {"1", "true", "yes", "on"}:
        run_yagpt_safe(result_dir)
        run_yandex_connectivity_test(result_dir)
    else:
        msg = "YandexGPT временно отключён. Для включения задайте ENABLE_YANDEXGPT=1."
        _write_text(os.path.join(result_dir, "Comprehensive_Legal_Review.txt"), msg)
        _write_text(os.path.join(result_dir, "YandexGPT_Connectivity_Test.txt"), msg)
        _write_json(os.path.join(result_dir, "YandexGPT_Connectivity_Log.json"), {"ok": False, "disabled": True})
        print(f"⏭ {msg}")

    # 2.2) Пинг OpenAI (результат или ошибка в OpenAI_Connectivity_Test.txt)
    run_openai_connectivity_test_safe(result_dir)

    # 2.x) Статус расходования токенов OpenAI (месяц/сегодня)
    write_openai_tokens_status(result_dir)

    # 2.3) Анализ материалов в OpenAI (результат или ошибка в Comprehensive_Legal_Review_OpenAI.txt)
    run_openai_materials_review_safe(result_dir)

    print("✅ Последовательный анализ завершён.")



# ===== НОВОЕ: финальный статус и маячок завершения =====
try:
    with open(_STATUS_PATH, "w", encoding="utf-8") as f:
        json.dump(
            {"state": "done", "stage": "finish", "ts": time.strftime("%Y-%m-%d %H:%M:%S")},
            f, ensure_ascii=False, indent=2
        )
except Exception:
    pass

try:
    open(os.path.join(result_dir, ".done"), "w", encoding="utf-8").close()
except Exception:
    pass

print("\n================ Clipchecker: DONE ==================\n", flush=True)
    # ===== /НОВОЕ =====


    # Асинхронный вызов Telegram бота (если нужно)
    # asyncio.run(recognize_faces_via_telegram_bot(faces_dir, telegram_faces_output))
