# -*- coding: utf-8 -*-
"""Detect primary and secondary advertisers in video frames."""

from __future__ import annotations

import html
import re
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageStat

try:
    from black_bars import find_source_frames_dir
    from cv_detection import detect_cv_objects_in_frames, normalize_cv_label
    from frame_safety import frame_second_label, iter_ocr_lines, load_ocr_log, normalize_text
    from price_checks import read_text_file
except ImportError:
    from .black_bars import find_source_frames_dir
    from .cv_detection import detect_cv_objects_in_frames, normalize_cv_label
    from .frame_safety import frame_second_label, iter_ocr_lines, load_ocr_log, normalize_text
    from .price_checks import read_text_file


BRAND_DATABASE_PATH = Path(
    r"C:\Users\PTambulatov\Desktop\PT\2.Расчеты\50. ХАКАТОН\16. список брендов для 2рд\1.список брендов для 2рд.xlsx"
)
BRAND_DATABASE_DIR = BRAND_DATABASE_PATH.parent
BRAND_DATABASE_FILENAME = BRAND_DATABASE_PATH.name
SECONDARY_BRANDS_SHEET = "Brands"
ASIAN_BRANDS_SHEET = "Asia"
STOPLIST_BRANDS_SHEET = "Stoplist"

BRAND_BEARING_CV_LABELS = [
    "brand logo",
    "logo",
    "product packaging",
    "package",
    "packaging",
    "product label",
    "label",
    "bottle label",
    "can label",
    "box label",
    "shopping bag",
    "store sign",
    "shop sign",
    "billboard",
    "poster",
    "shelf product",
    "retail product",
]


@dataclass(frozen=True)
class BrandRecord:
    name: str
    variants: tuple[str, ...]


def _normalize_for_match(text: str) -> str:
    normalized = normalize_text(str(text or "")).lower().replace("ё", "е")
    normalized = re.sub(r"[®™©]", " ", normalized)
    normalized = re.sub(r"[^0-9a-zа-я]+", " ", normalized, flags=re.IGNORECASE)
    return re.sub(r"\s+", " ", normalized).strip()


def _variant_key(text: str) -> str:
    return _normalize_for_match(text)


def _variant_pattern(variant: str) -> re.Pattern:
    words = [re.escape(part) for part in _normalize_for_match(variant).split() if part]
    if not words:
        return re.compile(r"a^")
    body = r"[\s\-_./]*".join(words)
    return re.compile(rf"(?<![0-9a-zа-я]){body}(?![0-9a-zа-я])", flags=re.IGNORECASE)


def _split_variants(name: str) -> list[str]:
    raw = str(name or "").strip()
    if not raw:
        return []
    variants = [raw]
    for match in re.finditer(r"\(([^)]+)\)", raw):
        variants.extend(part.strip() for part in re.split(r"[,;/|]", match.group(1)) if part.strip())
    variants.append(re.sub(r"\s*\([^)]*\)", "", raw).strip())
    output = []
    seen = set()
    for variant in variants:
        key = _variant_key(variant)
        if len(key.replace(" ", "")) < 3 or key in seen:
            continue
        seen.add(key)
        output.append(variant)
    return output


def load_brand_database_sheet(path: str | Path | None, sheet_name: str) -> list[BrandRecord]:
    workbook_path = Path(path) if path is not None else BRAND_DATABASE_PATH
    if not workbook_path.is_file():
        return []
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return []
        worksheet = workbook[sheet_name]
        records = []
        seen = set()
        for row in worksheet.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            name = str(row[0]).strip()
            key = _variant_key(name)
            if not key or key in seen:
                continue
            variants = tuple(_split_variants(name))
            if not variants:
                continue
            seen.add(key)
            records.append(BrandRecord(name=name, variants=variants))
        return records
    finally:
        workbook.close()


def load_secondary_brand_database(path: str | Path | None = None) -> list[BrandRecord]:
    return load_brand_database_sheet(path, SECONDARY_BRANDS_SHEET)


def load_asian_brand_database(path: str | Path | None = None) -> list[BrandRecord]:
    return load_brand_database_sheet(path, ASIAN_BRANDS_SHEET)


def load_stoplist_brand_database(path: str | Path | None = None) -> list[BrandRecord]:
    return load_brand_database_sheet(path, STOPLIST_BRANDS_SHEET)


def _compile_brand_patterns(records: list[BrandRecord]) -> list[tuple[BrandRecord, str, re.Pattern]]:
    patterns = []
    for record in records:
        for variant in record.variants:
            patterns.append((record, variant, _variant_pattern(variant)))
    return patterns


def _find_brand_mentions_in_text(text: str, patterns: list[tuple[BrandRecord, str, re.Pattern]]) -> list[dict]:
    normalized = _normalize_for_match(text)
    found = []
    seen = set()
    for record, variant, pattern in patterns:
        if not pattern.search(normalized):
            continue
        key = record.name.lower()
        if key in seen:
            continue
        seen.add(key)
        found.append({"brand": record.name, "variant": variant, "text": text})
    return found


def _sort_seconds(seconds: list[str]) -> list[str]:
    return sorted(
        set(seconds),
        key=lambda value: int(re.search(r"\d+", value).group(0)) if re.search(r"\d+", value) else 0,
    )


def analyze_brands_from_ocr(result_dir: str | Path, records: list[BrandRecord]) -> dict:
    base = Path(result_dir)
    ocr_path = base / "OCR_Log.json"
    patterns = _compile_brand_patterns(records)
    grouped: dict[str, dict] = {}
    if not ocr_path.is_file():
        text = read_text_file(base / "All_Frames_Text.txt")
        for mention in _find_brand_mentions_in_text(text, patterns):
            grouped.setdefault(
                mention["brand"],
                {"brand": mention["brand"], "variants": [], "seconds": [], "frames": [], "examples": [], "sources": []},
            )
            grouped[mention["brand"]]["variants"].append(mention["variant"])
            grouped[mention["brand"]]["sources"].append("text")
        return {"brand_mentions": _finalize_brand_mentions(grouped), "has_ocr": bool(text.strip())}

    for item in iter_ocr_lines(load_ocr_log(ocr_path)):
        text = normalize_text(item.get("text", ""))
        frame = item.get("frame", "")
        second = frame_second_label(frame)
        for mention in _find_brand_mentions_in_text(text, patterns):
            entry = grouped.setdefault(
                mention["brand"],
                {"brand": mention["brand"], "variants": [], "seconds": [], "frames": [], "examples": [], "sources": []},
            )
            entry["variants"].append(mention["variant"])
            entry["seconds"].append(second)
            entry["frames"].append(frame)
            entry["sources"].append("ocr")
            if len(entry["examples"]) < 3:
                entry["examples"].append(text)
    return {"brand_mentions": _finalize_brand_mentions(grouped), "has_ocr": True}


def _finalize_brand_mentions(grouped: dict[str, dict]) -> list[dict]:
    output = []
    for item in grouped.values():
        item["variants"] = sorted(set(item.get("variants", [])), key=str.lower)
        item["seconds"] = _sort_seconds(item.get("seconds", []))
        item["frames"] = sorted(set(item.get("frames", [])))
        item["sources"] = sorted(set(item.get("sources", [])))
        item["duration_sec"] = len(item["seconds"])
        output.append(item)
    return sorted(output, key=lambda item: (-item["duration_sec"], item["brand"].lower()))


def analyze_brands_from_documents(result_dir: str | Path, records: list[BrandRecord]) -> list[dict]:
    text = read_text_file(Path(result_dir) / "Documents_Texts.txt")
    if not text.strip():
        return []
    patterns = _compile_brand_patterns(records)
    grouped: dict[str, dict] = {}
    for mention in _find_brand_mentions_in_text(text, patterns):
        entry = grouped.setdefault(mention["brand"], {"brand": mention["brand"], "variants": [], "count": 0})
        entry["variants"].append(mention["variant"])
        entry["count"] += len(_variant_pattern(mention["variant"]).findall(_normalize_for_match(text)))
    output = []
    for item in grouped.values():
        item["variants"] = sorted(set(item["variants"]), key=str.lower)
        output.append(item)
    return sorted(output, key=lambda item: (-item["count"], item["brand"].lower()))


def _all_cv_labels(records: list[BrandRecord]) -> list[str]:
    labels = list(BRAND_BEARING_CV_LABELS)
    seen = {normalize_cv_label(label) for label in labels}
    for record in records:
        for variant in record.variants:
            label = str(variant).strip()
            key = normalize_cv_label(label)
            if key and key not in seen:
                seen.add(key)
                labels.append(label)
    return labels


def analyze_secondary_advertisers_from_cv(result_dir: str | Path, records: list[BrandRecord]) -> dict:
    frames_dir = find_source_frames_dir(result_dir)
    if frames_dir is None:
        return {
            "cv_enabled": False,
            "cv_model_path": "",
            "cv_error": "Кадры для CV-проверки не найдены.",
            "cv_brand_mentions": [],
            "cv_brand_bearing_objects": [],
        }
    cv_result = detect_cv_objects_in_frames(frames_dir, labels=_all_cv_labels(records))
    variant_to_brand = {}
    for record in records:
        for variant in record.variants:
            variant_to_brand[normalize_cv_label(variant)] = record.name

    brand_grouped: dict[str, dict] = {}
    brand_objects = []
    generic_labels = {normalize_cv_label(label) for label in BRAND_BEARING_CV_LABELS}
    for detection in cv_result["detections"]:
        label_key = normalize_cv_label(detection.get("raw_label") or detection.get("label", ""))
        if label_key in variant_to_brand:
            brand = variant_to_brand[label_key]
            entry = brand_grouped.setdefault(
                brand,
                {"brand": brand, "seconds": [], "frames": [], "sources": [], "cv_labels": []},
            )
            entry["seconds"].append(detection.get("second", ""))
            entry["frames"].append(detection.get("frame", ""))
            entry["sources"].append("cv")
            entry["cv_labels"].append(detection.get("raw_label") or detection.get("label"))
        elif label_key in generic_labels:
            brand_objects.append(detection)

    mentions = []
    for item in brand_grouped.values():
        item["seconds"] = _sort_seconds([s for s in item["seconds"] if s])
        item["frames"] = sorted({frame for frame in item["frames"] if frame})
        item["sources"] = sorted(set(item["sources"]))
        item["cv_labels"] = sorted({label for label in item["cv_labels"] if label}, key=str.lower)
        item["duration_sec"] = len(item["seconds"])
        mentions.append(item)

    return {
        "cv_enabled": cv_result["enabled"],
        "cv_model_path": cv_result["model_path"],
        "cv_error": cv_result["error"],
        "cv_brand_mentions": sorted(mentions, key=lambda item: (-item["duration_sec"], item["brand"].lower())),
        "cv_brand_bearing_objects": brand_objects,
    }


def _merge_brand_mentions(*groups: list[dict]) -> list[dict]:
    merged: dict[str, dict] = {}
    for group in groups:
        for item in group:
            brand = item["brand"]
            entry = merged.setdefault(
                brand,
                {"brand": brand, "variants": [], "seconds": [], "frames": [], "examples": [], "sources": [], "cv_labels": []},
            )
            for key in ("variants", "seconds", "frames", "examples", "sources", "cv_labels"):
                entry[key].extend(item.get(key, []))
    return _finalize_brand_mentions(merged)


def _choose_primary_advertiser(video_mentions: list[dict], document_mentions: list[dict]) -> str | None:
    if document_mentions:
        return document_mentions[0]["brand"]
    if video_mentions:
        return video_mentions[0]["brand"]
    return None


def _format_brand_html(brand: str, color: str = "#16a34a") -> str:
    return f'<strong style="color:{color};font-weight:800">{html.escape(brand)}</strong>'


def _format_secondary_html(mentions: list[dict]) -> str:
    parts = []
    for item in mentions:
        duration = item.get("duration_sec", 0)
        seconds = ", ".join(item.get("seconds", [])[:12])
        suffix = f" - {duration} сек." if duration else ""
        if seconds:
            suffix += f" ({html.escape(seconds)})"
        parts.append(f'{_format_brand_html(item["brand"], "#dc2626")}{suffix}')
    return "; ".join(parts)


def _brand_item_keys(item: dict) -> set[str]:
    values = [item.get("brand", "")]
    values.extend(item.get("variants", []))
    values.extend(item.get("cv_labels", []))
    return {key for value in values if (key := _variant_key(value))}


def _stoplist_key_to_brand(records: list[BrandRecord]) -> dict[str, str]:
    key_to_brand = {}
    for record in records:
        for value in (record.name, *record.variants):
            key = _variant_key(value)
            if key:
                key_to_brand[key] = record.name
    return key_to_brand


def _match_mentions_to_stoplist(mentions: list[dict], stoplist_records: list[BrandRecord]) -> list[dict]:
    stoplist_keys = _stoplist_key_to_brand(stoplist_records)
    matches = []
    for item in mentions:
        matched_names = sorted({stoplist_keys[key] for key in _brand_item_keys(item) if key in stoplist_keys}, key=str.lower)
        if not matched_names:
            continue
        matched = dict(item)
        matched["stoplist_brands"] = matched_names
        matches.append(matched)
    return matches


def _iter_frame_files(frames_dir: Path) -> list[Path]:
    if not frames_dir.is_dir():
        return []
    return sorted(
        [
            path
            for path in frames_dir.iterdir()
            if path.is_file() and path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp"}
        ],
        key=lambda path: frame_second_label(path.name),
    )


def _frame_lookup(result_dir: str | Path) -> dict[str, Path]:
    frames_dir = find_source_frames_dir(result_dir)
    if frames_dir is None:
        return {}
    return {path.name: path for path in _iter_frame_files(frames_dir)}


def _style_vector(path: Path) -> list[float] | None:
    try:
        with Image.open(path) as img:
            rgb = img.convert("RGB").resize((64, 64))
            stat = ImageStat.Stat(rgb)
            means = [value / 255.0 for value in stat.mean]
            stddev = [value / 128.0 for value in stat.stddev]
            histogram = rgb.histogram()
    except OSError:
        return None

    features: list[float] = []
    for channel in range(3):
        start = channel * 256
        channel_hist = histogram[start : start + 256]
        total = float(sum(channel_hist)) or 1.0
        for bin_start in range(0, 256, 32):
            features.append(sum(channel_hist[bin_start : bin_start + 32]) / total)
    features.extend(means)
    features.extend(stddev)
    return features


def _mean_vector(vectors: list[list[float]]) -> list[float] | None:
    if not vectors:
        return None
    return [sum(vector[idx] for vector in vectors) / len(vectors) for idx in range(len(vectors[0]))]


def _style_distance(left: list[float], right: list[float]) -> float:
    if not left or not right or len(left) != len(right):
        return 1.0
    return sum(abs(a - b) for a, b in zip(left, right)) / len(left)


def _format_style_issues_html(issues: list[dict]) -> str:
    parts = []
    for issue in issues:
        brand = _format_brand_html(issue["brand"], "#dc2626")
        frames = ", ".join(html.escape(frame) for frame in issue.get("frames", [])[:8])
        score = issue.get("max_distance", 0)
        suffix = f" - отличие стиля {score:.2f}"
        if frames:
            suffix += f" ({frames})"
        parts.append(f"{brand}{suffix}")
    return "; ".join(parts)


def _has_materials(result_dir: Path) -> bool:
    return (
        (result_dir / "OCR_Log.json").is_file()
        or (result_dir / "All_Frames_Text.txt").is_file()
        or (result_dir / "frames").is_dir()
        or (result_dir / "frames_pdf_original").is_dir()
    )


def evaluate_secondary_advertisers(result_dir: str | Path) -> dict:
    base = Path(result_dir)
    records = load_secondary_brand_database()
    if not records:
        return {
            "status": "pending",
            "message": f"База второстепенных рекламодателей не найдена: {BRAND_DATABASE_PATH}",
            "brand_database_path": str(BRAND_DATABASE_PATH),
            "brand_database_count": 0,
        }

    ocr_analysis = analyze_brands_from_ocr(base, records)
    document_mentions = analyze_brands_from_documents(base, records)
    cv_analysis = analyze_secondary_advertisers_from_cv(base, records)
    video_mentions = _merge_brand_mentions(ocr_analysis["brand_mentions"], cv_analysis["cv_brand_mentions"])
    primary = _choose_primary_advertiser(video_mentions, document_mentions)
    secondary = [item for item in video_mentions if item["brand"] != primary]

    common = {
        "brand_database_path": str(BRAND_DATABASE_PATH),
        "brand_database_count": len(records),
        "primary_advertiser": primary,
        "video_brand_mentions": video_mentions,
        "document_brand_mentions": document_mentions,
        **cv_analysis,
    }

    if not video_mentions:
        if not _has_materials(base):
            return {
                "status": "pending",
                "message": "Проверка рекламодателей будет выполнена после извлечения кадров, OCR и CV-анализа.",
                **common,
            }
        if cv_analysis.get("cv_brand_bearing_objects"):
            count = len(cv_analysis["cv_brand_bearing_objects"])
            return {
                "status": "warning",
                "message": f"Найдены упаковки, этикетки или логотипные зоны без распознанного бренда: {count}. Требуется ручная проверка второстепенных рекламодателей.",
                **common,
            }
        return {
            "status": "warning",
            "message": "Рекламодатели в видеоряде не распознаны по базе второстепенных рекламодателей.",
            **common,
        }

    if not secondary:
        primary_name = primary or video_mentions[0]["brand"]
        return {
            "status": "pass",
            "message": f"В видеоряде распознан один основной рекламодатель: {primary_name}.",
            "message_html": f"В видеоряде распознан один основной рекламодатель: {_format_brand_html(primary_name)}.",
            **common,
        }

    return {
        "status": "fail",
        "message": "Обнаружены второстепенные рекламодатели: "
        + "; ".join(f"{item['brand']} - {item.get('duration_sec', 0)} сек." for item in secondary)
        + ".",
        "message_html": "Обнаружены второстепенные рекламодатели: " + _format_secondary_html(secondary) + ".",
        "secondary_advertisers": secondary,
        **common,
    }


def evaluate_asian_brands(result_dir: str | Path) -> dict:
    base = Path(result_dir)
    records = load_asian_brand_database()
    if not records:
        return {
            "status": "pending",
            "message": f"Список азиатских брендов не найден во вкладке {ASIAN_BRANDS_SHEET}: {BRAND_DATABASE_PATH}",
            "brand_database_path": str(BRAND_DATABASE_PATH),
            "brand_database_sheet": ASIAN_BRANDS_SHEET,
            "brand_database_count": 0,
        }

    ocr_analysis = analyze_brands_from_ocr(base, records)
    cv_analysis = analyze_secondary_advertisers_from_cv(base, records)
    brand_mentions = _merge_brand_mentions(ocr_analysis["brand_mentions"], cv_analysis["cv_brand_mentions"])
    common = {
        "brand_database_path": str(BRAND_DATABASE_PATH),
        "brand_database_sheet": ASIAN_BRANDS_SHEET,
        "brand_database_count": len(records),
        "asian_brand_mentions": brand_mentions,
        **cv_analysis,
    }

    if not brand_mentions:
        if not _has_materials(base):
            return {
                "status": "pending",
                "message": "Проверка азиатских брендов будет выполнена после извлечения кадров, OCR и CV-анализа.",
                **common,
            }
        return {
            "status": "pass",
            "message": "Азиатских брендов не обнаружено",
            **common,
        }

    return {
        "status": "warning",
        "message": "Обнаружены азиатские бренды: "
        + "; ".join(f"{item['brand']} - {item.get('duration_sec', 0)} сек." for item in brand_mentions)
        + ".",
        "message_html": "Обнаружены азиатские бренды: " + _format_secondary_html(brand_mentions) + ".",
        **common,
    }


def evaluate_secondary_advertiser_stoplist(result_dir: str | Path) -> dict:
    stoplist_records = load_stoplist_brand_database()
    if not stoplist_records:
        return {
            "status": "pending",
            "message": f"Стоп-лист брендов не найден во вкладке {STOPLIST_BRANDS_SHEET}: {BRAND_DATABASE_PATH}",
            "brand_database_path": str(BRAND_DATABASE_PATH),
            "brand_database_sheet": STOPLIST_BRANDS_SHEET,
            "brand_database_count": 0,
        }

    secondary_result = evaluate_secondary_advertisers(result_dir)
    secondary = secondary_result.get("secondary_advertisers") or []
    common = {
        "brand_database_path": str(BRAND_DATABASE_PATH),
        "brand_database_sheet": STOPLIST_BRANDS_SHEET,
        "brand_database_count": len(stoplist_records),
        "secondary_advertisers": secondary,
        "secondary_advertiser_result": secondary_result,
    }

    if not secondary:
        if secondary_result.get("status") in {"pending", "warning"}:
            return {
                "status": secondary_result.get("status", "pending"),
                "message": "Стоп-лист второстепенных рекламодателей не проверен: "
                + secondary_result.get("message", "второстепенные рекламодатели не определены."),
                **common,
            }
        return {
            "status": "pass",
            "message": "Второстепенные рекламодатели не обнаружены; стоп-лист не применим.",
            **common,
        }

    matches = _match_mentions_to_stoplist(secondary, stoplist_records)
    if matches:
        return {
            "status": "fail",
            "message": "Второстепенный рекламодатель входит в стоп-лист: "
            + "; ".join(item["brand"] for item in matches)
            + ".",
            "message_html": "Второстепенный рекламодатель входит в стоп-лист: " + _format_secondary_html(matches) + ".",
            "stoplist_matches": matches,
            **common,
        }

    return {
        "status": "pass",
        "message": "Второстепенные рекламодатели не входят в стоп-лист: "
        + "; ".join(item["brand"] for item in secondary)
        + ".",
        **common,
    }


def evaluate_secondary_advertiser_style(result_dir: str | Path) -> dict:
    secondary_result = evaluate_secondary_advertisers(result_dir)
    secondary = secondary_result.get("secondary_advertisers") or []
    common = {
        "secondary_advertiser_result": secondary_result,
        "primary_advertiser": secondary_result.get("primary_advertiser"),
        "secondary_advertisers": secondary,
        "style_distance_threshold": 0.23,
    }

    if not secondary:
        if secondary_result.get("status") in {"pending", "warning"}:
            return {
                "status": secondary_result.get("status", "pending"),
                "message": "Стилистика ролика не проверена: "
                + secondary_result.get("message", "второстепенные рекламодатели не определены."),
                **common,
            }
        return {
            "status": "pass",
            "message": "В ролике не обнаружено больше одного рекламодателя; проверка стилистики не применима.",
            **common,
        }

    primary = secondary_result.get("primary_advertiser")
    video_mentions = secondary_result.get("video_brand_mentions") or []
    primary_mentions = [item for item in video_mentions if item.get("brand") == primary]
    primary_frames = sorted({frame for item in primary_mentions for frame in item.get("frames", []) if frame})
    secondary_frames = sorted({frame for item in secondary for frame in item.get("frames", []) if frame})

    lookup = _frame_lookup(result_dir)
    if not lookup:
        return {
            "status": "pending",
            "message": "Кадры для проверки стилистики не найдены.",
            **common,
        }
    if not primary_frames:
        secondary_frame_set = set(secondary_frames)
        primary_frames = [name for name in lookup if name not in secondary_frame_set]
    if not primary_frames:
        return {
            "status": "pending",
            "message": "Недостаточно кадров основного рекламодателя для сравнения стилистики.",
            **common,
        }

    primary_vectors = [_style_vector(lookup[frame]) for frame in primary_frames if frame in lookup]
    primary_vectors = [vector for vector in primary_vectors if vector is not None]
    primary_profile = _mean_vector(primary_vectors)
    if primary_profile is None:
        return {
            "status": "pending",
            "message": "Не удалось рассчитать визуальный профиль кадров основного рекламодателя.",
            **common,
        }

    threshold = common["style_distance_threshold"]
    issues = []
    brand_style_results = []
    for item in secondary:
        frame_distances = []
        for frame in item.get("frames", []):
            path = lookup.get(frame)
            vector = _style_vector(path) if path is not None else None
            if vector is None:
                continue
            frame_distances.append({"frame": frame, "distance": _style_distance(primary_profile, vector)})
        if not frame_distances:
            continue
        max_distance = max(value["distance"] for value in frame_distances)
        avg_distance = sum(value["distance"] for value in frame_distances) / len(frame_distances)
        result = {
            "brand": item["brand"],
            "frames": [value["frame"] for value in frame_distances],
            "max_distance": round(max_distance, 3),
            "avg_distance": round(avg_distance, 3),
        }
        brand_style_results.append(result)
        if max_distance > threshold:
            issues.append(result)

    common["style_results"] = brand_style_results
    if not brand_style_results:
        return {
            "status": "pending",
            "message": "Недостаточно кадров второстепенных рекламодателей для сравнения стилистики.",
            **common,
        }

    if issues:
        return {
            "status": "fail",
            "message": "Кадры со второстепенными рекламодателями заметно отличаются от стилистики основного рекламодателя: "
            + "; ".join(f"{item['brand']} - отличие стиля {item['max_distance']:.2f}" for item in issues)
            + ".",
            "message_html": "Кадры со второстепенными рекламодателями заметно отличаются от стилистики основного рекламодателя: "
            + _format_style_issues_html(issues)
            + ".",
            **common,
        }

    return {
        "status": "pass",
        "message": "Кадры со второстепенными рекламодателями не имеют кардинального отличия от стилистики основного рекламодателя.",
        **common,
    }
