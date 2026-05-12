# -*- coding: utf-8 -*-
"""Read the advertiser category approval checklist from the Excel map."""

from __future__ import annotations

from collections import OrderedDict
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


APPROVAL_MAP_PATH = Path(
    r"C:\Users\PTambulatov\Desktop\PT\2.Расчеты\50. ХАКАТОН\14. карта согласования рм\1. Карта согласования РМ.xlsx"
)
APPROVAL_SHEET_NAME = "РМ"
REQUIRED_BLOCK_ORDER = [
    "Видеоряд",
    "Набивка",
    "Документы",
    "Озвучка",
    "Требования НРА",
    "Требования ЗоР",
]


@dataclass(frozen=True)
class ApprovalItem:
    id: str
    block: str
    number: str
    text: str


def _norm(value) -> str:
    return str(value or "").replace("\xa0", " ").strip()


def _is_selected(value) -> bool:
    if value is None:
        return False
    if isinstance(value, (int, float)):
        return int(value) == 1
    return _norm(value).lower() in {"1", "1.0", "да", "yes", "true", "x", "+"}


def _find_header_row(ws) -> tuple[int, dict[str, int], list[tuple[str, int]]]:
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        cells = [_norm(ws.cell(row_idx, col).value) for col in range(1, ws.max_column + 1)]
        lower = [c.lower() for c in cells]
        if "блок проверки" not in lower or "проверка" not in lower:
            continue

        block_col = lower.index("блок проверки") + 1
        check_col = lower.index("проверка") + 1
        id_col = lower.index("id") + 1 if "id" in lower else None
        number_col = lower.index("номер") + 1 if "номер" in lower else None
        categories: list[tuple[str, int]] = []
        for col in range(check_col + 1, ws.max_column + 1):
            name = _norm(ws.cell(row_idx, col).value)
            if name:
                categories.append((name, col))
        return row_idx, {"block": block_col, "id": id_col or 0, "number": number_col or 0, "check": check_col}, categories

    raise ValueError("В листе РМ не найдена строка заголовков с колонками 'Блок проверки' и 'Проверка'")


def load_approval_categories(path: Path | str = APPROVAL_MAP_PATH) -> list[str]:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        ws = wb[APPROVAL_SHEET_NAME]
        _, _, categories = _find_header_row(ws)
        return [name for name, _ in categories]
    finally:
        wb.close()


def load_approval_checklist(category: str, path: Path | str = APPROVAL_MAP_PATH) -> OrderedDict[str, list[ApprovalItem]]:
    workbook_path = Path(path)
    wb = load_workbook(workbook_path, data_only=True, read_only=True)
    try:
        ws = wb[APPROVAL_SHEET_NAME]
        header_row, columns, categories = _find_header_row(ws)
        category_map = {name: col for name, col in categories}
        if category not in category_map:
            raise ValueError(f"Категория '{category}' не найдена в карте согласования")

        selected_col = category_map[category]
        blocks: OrderedDict[str, list[ApprovalItem]] = OrderedDict((block, []) for block in REQUIRED_BLOCK_ORDER)

        for row_idx in range(header_row + 1, ws.max_row + 1):
            if not _is_selected(ws.cell(row_idx, selected_col).value):
                continue

            block = _norm(ws.cell(row_idx, columns["block"]).value)
            text = _norm(ws.cell(row_idx, columns["check"]).value)
            if not block or not text:
                continue

            item_id = _norm(ws.cell(row_idx, columns["id"]).value) if columns["id"] else ""
            number = _norm(ws.cell(row_idx, columns["number"]).value) if columns["number"] else ""
            if block not in blocks:
                blocks[block] = []
            blocks[block].append(ApprovalItem(id=item_id, block=block, number=number, text=text))

        return OrderedDict((block, items) for block, items in blocks.items() if items)
    finally:
        wb.close()


def build_approval_view_model(category: str | None, path: Path | str = APPROVAL_MAP_PATH) -> dict:
    try:
        categories = load_approval_categories(path)
        selected = category if category in categories else (categories[0] if categories else "")
        blocks = load_approval_checklist(selected, path) if selected else OrderedDict()
        return {
            "ok": True,
            "categories": categories,
            "selected_category": selected,
            "blocks": [
                {
                    "name": block,
                    "items": [
                        {"id": item.id, "number": item.number, "text": item.text}
                        for item in items
                    ],
                }
                for block, items in blocks.items()
            ],
            "error": "",
        }
    except Exception as exc:
        return {
            "ok": False,
            "categories": [],
            "selected_category": category or "",
            "blocks": [],
            "error": str(exc),
        }
