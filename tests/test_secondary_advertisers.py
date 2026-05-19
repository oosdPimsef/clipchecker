# -*- coding: utf-8 -*-

import json
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from openpyxl import Workbook
from PIL import Image

from app.approval_checks import evaluate_approval_view_model
from app.secondary_advertisers import (
    BRAND_BEARING_CV_LABELS,
    evaluate_asian_brands,
    evaluate_secondary_advertiser_stoplist,
    evaluate_secondary_advertiser_style,
    evaluate_secondary_advertisers,
    load_asian_brand_database,
    load_secondary_brand_database,
    load_stoplist_brand_database,
)


def make_ocr_line(text: str, box: tuple[int, int, int, int] = (120, 80, 700, 140)) -> dict:
    x1, y1, x2, y2 = box
    vertices = [
        {"x": str(x1), "y": str(y1)},
        {"x": str(x1), "y": str(y2)},
        {"x": str(x2), "y": str(y2)},
        {"x": str(x2), "y": str(y1)},
    ]
    return {
        "boundingBox": {"vertices": vertices},
        "words": [{"text": word, "boundingBox": {"vertices": vertices}} for word in text.split()],
    }


def make_ocr_log(frame_lines: dict[str, list[str]]) -> dict:
    data = {}
    for frame, lines in frame_lines.items():
        data[frame] = {
            "results": [
                {
                    "results": [
                        {
                            "textDetection": {
                                "pages": [{"blocks": [{"lines": [make_ocr_line(line) for line in lines]}]}]
                            }
                        }
                    ]
                }
            ]
        }
    return data


def make_brand_database(
    path: Path,
    brands: list[str],
    asia_brands: list[str] | None = None,
    stoplist_brands: list[str] | None = None,
) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Brands"
    for row, brand in enumerate(brands, start=1):
        worksheet.cell(row=row, column=1, value=brand)
    if asia_brands is not None:
        asia = workbook.create_sheet("Asia")
        for row, brand in enumerate(asia_brands, start=1):
            asia.cell(row=row, column=1, value=brand)
    if stoplist_brands is not None:
        stoplist = workbook.create_sheet("Stoplist")
        for row, brand in enumerate(stoplist_brands, start=1):
            stoplist.cell(row=row, column=1, value=brand)
    workbook.save(path)


def make_result_dir(
    ocr_log: dict | None = None,
    documents_text: str = "",
    frame_colors: dict[str, tuple[int, int, int] | str] | None = None,
):
    tmp = tempfile.TemporaryDirectory()
    base = Path(tmp.name)
    frames = base / "frames"
    frames.mkdir()
    frame_names = list(ocr_log or {"frame_001.jpg": []})
    for frame_name in frame_names:
        color = (frame_colors or {}).get(frame_name, "white")
        Image.new("RGB", (1000, 500), color).save(frames / frame_name)
    if ocr_log is not None:
        (base / "OCR_Log.json").write_text(json.dumps(ocr_log, ensure_ascii=False), encoding="utf-8")
    if documents_text:
        (base / "Documents_Texts.txt").write_text(documents_text, encoding="utf-8")
    return tmp, base


class SecondaryAdvertisersTests(unittest.TestCase):
    def setUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.brand_path = Path(self.tmp.name) / "brands.xlsx"
        make_brand_database(
            self.brand_path,
            ["Основной бренд", "PepsiCo", "X5 RETAIL GROUP", "АВТОВАЗ (LADA)"],
            ["XIAOMI", "Дахуа", "70mai"],
            ["PepsiCo", "Apple"],
        )
        self.brand_patch = patch("app.secondary_advertisers.BRAND_DATABASE_PATH", self.brand_path)
        self.cv_patch = patch(
            "app.secondary_advertisers.analyze_secondary_advertisers_from_cv",
            return_value={
                "cv_enabled": False,
                "cv_model_path": "",
                "cv_error": "YOLO test disabled",
                "cv_brand_mentions": [],
                "cv_brand_bearing_objects": [],
            },
        )
        self.brand_patch.start()
        self.cv_patch.start()

    def tearDown(self):
        self.cv_patch.stop()
        self.brand_patch.stop()
        self.tmp.cleanup()

    def test_loads_brand_database_and_parentheses_variants(self):
        records = load_secondary_brand_database(self.brand_path)
        lada = [record for record in records if record.name == "АВТОВАЗ (LADA)"][0]
        self.assertIn("LADA", lada.variants)
        self.assertIn("АВТОВАЗ", lada.variants)

    def test_brand_database_is_read_strictly_from_brands_sheet(self):
        workbook = Workbook()
        workbook.active.title = "Other"
        workbook.active["A1"] = "PepsiCo"
        strict_path = Path(self.tmp.name) / "strict_brands.xlsx"
        workbook.save(strict_path)

        records = load_secondary_brand_database(strict_path)

        self.assertEqual(records, [])

    def test_asian_brand_database_is_read_strictly_from_asia_sheet(self):
        records = load_asian_brand_database(self.brand_path)
        self.assertEqual([record.name for record in records], ["XIAOMI", "Дахуа", "70mai"])

        workbook = Workbook()
        workbook.active.title = "Brands"
        workbook.active["A1"] = "XIAOMI"
        no_asia_path = Path(self.tmp.name) / "no_asia.xlsx"
        workbook.save(no_asia_path)

        self.assertEqual(load_asian_brand_database(no_asia_path), [])

    def test_stoplist_brand_database_is_read_strictly_from_stoplist_sheet(self):
        records = load_stoplist_brand_database(self.brand_path)
        self.assertEqual([record.name for record in records], ["PepsiCo", "Apple"])

        workbook = Workbook()
        workbook.active.title = "Brands"
        workbook.active["A1"] = "PepsiCo"
        no_stoplist_path = Path(self.tmp.name) / "no_stoplist.xlsx"
        workbook.save(no_stoplist_path)

        self.assertEqual(load_stoplist_brand_database(no_stoplist_path), [])

    def test_passes_when_only_primary_advertiser_is_found(self):
        tmp, base = make_result_dir(
            make_ocr_log({"frame_001.jpg": ["Основной бренд"], "frame_002.jpg": ["Основной бренд"]}),
            documents_text="Рекламодатель: Основной бренд",
        )
        try:
            result = evaluate_secondary_advertisers(base)
        finally:
            tmp.cleanup()

        self.assertEqual(result["status"], "pass")
        self.assertEqual(result["primary_advertiser"], "Основной бренд")
        self.assertIn('style="color:#16a34a;font-weight:800"', result["message_html"])

    def test_fails_and_counts_seconds_for_secondary_advertisers(self):
        tmp, base = make_result_dir(
            make_ocr_log(
                {
                    "frame_001.jpg": ["Основной бренд"],
                    "frame_002.jpg": ["PepsiCo"],
                    "frame_003.jpg": ["PepsiCo"],
                }
            ),
            documents_text="Рекламодатель: Основной бренд",
        )
        try:
            result = evaluate_secondary_advertisers(base)
        finally:
            tmp.cleanup()

        self.assertEqual(result["status"], "fail")
        self.assertEqual(result["primary_advertiser"], "Основной бренд")
        self.assertEqual(result["secondary_advertisers"][0]["brand"], "PepsiCo")
        self.assertEqual(result["secondary_advertisers"][0]["duration_sec"], 2)
        self.assertIn('style="color:#dc2626;font-weight:800"', result["message_html"])

    def test_cv_brand_detection_can_find_secondary_without_ocr(self):
        tmp, base = make_result_dir(make_ocr_log({"frame_001.jpg": ["Основной бренд"]}), documents_text="Основной бренд")
        fake_cv = {
            "cv_enabled": True,
            "cv_model_path": "model.pt",
            "cv_error": "",
            "cv_brand_mentions": [
                {
                    "brand": "PepsiCo",
                    "seconds": ["1сек."],
                    "frames": ["frame_001.jpg"],
                    "sources": ["cv"],
                    "cv_labels": ["PepsiCo"],
                    "duration_sec": 1,
                }
            ],
            "cv_brand_bearing_objects": [],
        }
        try:
            with patch("app.secondary_advertisers.analyze_secondary_advertisers_from_cv", return_value=fake_cv):
                result = evaluate_secondary_advertisers(base)
        finally:
            tmp.cleanup()

        self.assertEqual(result["status"], "fail")
        self.assertEqual(result["secondary_advertisers"][0]["brand"], "PepsiCo")

    def test_warns_when_packaging_or_label_zone_is_unrecognized(self):
        tmp, base = make_result_dir(make_ocr_log({"frame_001.jpg": ["нейтральный текст"]}))
        fake_cv = {
            "cv_enabled": True,
            "cv_model_path": "model.pt",
            "cv_error": "",
            "cv_brand_mentions": [],
            "cv_brand_bearing_objects": [
                {"label": "product packaging", "raw_label": "product packaging", "frame": "frame_001.jpg", "second": "1сек.", "confidence": 0.8}
            ],
        }
        try:
            with patch("app.secondary_advertisers.analyze_secondary_advertisers_from_cv", return_value=fake_cv):
                result = evaluate_secondary_advertisers(base)
        finally:
            tmp.cleanup()

        self.assertEqual(result["status"], "warning")
        self.assertIn("упаковки", result["message"])

    def test_evaluates_item_id_220_inside_view_model(self):
        tmp, base = make_result_dir(
            make_ocr_log({"frame_001.jpg": ["Основной бренд"], "frame_002.jpg": ["X5 RETAIL GROUP"]}),
            documents_text="Основной бренд",
        )
        view_model = {
            "ok": True,
            "blocks": [{"name": "Видеоряд", "items": [{"id": "220", "number": "220", "text": "2rd"}]}],
        }
        try:
            evaluated = evaluate_approval_view_model(view_model, base)
        finally:
            tmp.cleanup()

        item = evaluated["blocks"][0]["items"][0]
        self.assertEqual(item["status"], "fail")
        self.assertIn("X5 RETAIL GROUP", item["message"])

    def test_secondary_advertiser_stoplist_fails_for_stoplisted_secondary(self):
        tmp, base = make_result_dir(
            make_ocr_log({"frame_001.jpg": ["Основной бренд"], "frame_002.jpg": ["PepsiCo"]}),
            documents_text="Основной бренд",
        )
        try:
            result = evaluate_secondary_advertiser_stoplist(base)
        finally:
            tmp.cleanup()

        self.assertEqual(result["status"], "fail")
        self.assertEqual(result["stoplist_matches"][0]["brand"], "PepsiCo")
        self.assertIn('style="color:#dc2626;font-weight:800"', result["message_html"])

    def test_secondary_advertiser_stoplist_passes_for_non_stoplisted_secondary(self):
        tmp, base = make_result_dir(
            make_ocr_log({"frame_001.jpg": ["Основной бренд"], "frame_002.jpg": ["X5 RETAIL GROUP"]}),
            documents_text="Основной бренд",
        )
        try:
            result = evaluate_secondary_advertiser_stoplist(base)
        finally:
            tmp.cleanup()

        self.assertEqual(result["status"], "pass")
        self.assertIn("X5 RETAIL GROUP", result["message"])

    def test_secondary_advertiser_stoplist_passes_when_no_secondary_advertiser(self):
        tmp, base = make_result_dir(
            make_ocr_log({"frame_001.jpg": ["Основной бренд"]}),
            documents_text="Основной бренд",
        )
        try:
            result = evaluate_secondary_advertiser_stoplist(base)
        finally:
            tmp.cleanup()

        self.assertEqual(result["status"], "pass")
        self.assertIn("не применим", result["message"])

    def test_evaluates_item_id_223_inside_view_model(self):
        tmp, base = make_result_dir(
            make_ocr_log({"frame_001.jpg": ["Основной бренд"], "frame_002.jpg": ["PepsiCo"]}),
            documents_text="Основной бренд",
        )
        view_model = {
            "ok": True,
            "blocks": [{"name": "Видеоряд", "items": [{"id": "223", "number": "223", "text": "stoplist"}]}],
        }
        try:
            evaluated = evaluate_approval_view_model(view_model, base)
        finally:
            tmp.cleanup()

        item = evaluated["blocks"][0]["items"][0]
        self.assertEqual(item["status"], "fail")
        self.assertIn("PepsiCo", item["message"])

    def test_secondary_advertiser_style_passes_when_frames_are_similar(self):
        tmp, base = make_result_dir(
            make_ocr_log({"frame_001.jpg": ["Основной бренд"], "frame_002.jpg": ["X5 RETAIL GROUP"]}),
            documents_text="Основной бренд",
            frame_colors={"frame_001.jpg": (245, 245, 245), "frame_002.jpg": (235, 235, 235)},
        )
        try:
            result = evaluate_secondary_advertiser_style(base)
        finally:
            tmp.cleanup()

        self.assertEqual(result["status"], "pass")
        self.assertIn("не имеют кардинального отличия", result["message"])

    def test_secondary_advertiser_style_fails_when_frames_change_style(self):
        tmp, base = make_result_dir(
            make_ocr_log({"frame_001.jpg": ["Основной бренд"], "frame_002.jpg": ["X5 RETAIL GROUP"]}),
            documents_text="Основной бренд",
            frame_colors={"frame_001.jpg": (255, 255, 255), "frame_002.jpg": (0, 0, 0)},
        )
        try:
            result = evaluate_secondary_advertiser_style(base)
        finally:
            tmp.cleanup()

        self.assertEqual(result["status"], "fail")
        self.assertIn("X5 RETAIL GROUP", result["message"])
        self.assertIn('style="color:#dc2626;font-weight:800"', result["message_html"])

    def test_evaluates_item_id_225_inside_view_model(self):
        tmp, base = make_result_dir(
            make_ocr_log({"frame_001.jpg": ["Основной бренд"], "frame_002.jpg": ["X5 RETAIL GROUP"]}),
            documents_text="Основной бренд",
            frame_colors={"frame_001.jpg": (255, 255, 255), "frame_002.jpg": (0, 0, 0)},
        )
        view_model = {
            "ok": True,
            "blocks": [{"name": "Видеоряд", "items": [{"id": "225", "number": "225", "text": "style"}]}],
        }
        try:
            evaluated = evaluate_approval_view_model(view_model, base)
        finally:
            tmp.cleanup()

        item = evaluated["blocks"][0]["items"][0]
        self.assertEqual(item["status"], "fail")
        self.assertIn("X5 RETAIL GROUP", item["message"])

    def test_asian_brands_warns_and_formats_brand_names(self):
        tmp, base = make_result_dir(make_ocr_log({"frame_001.jpg": ["XIAOMI"], "frame_002.jpg": ["Дахуа"]}))
        try:
            result = evaluate_asian_brands(base)
        finally:
            tmp.cleanup()

        self.assertEqual(result["status"], "warning")
        self.assertIn("XIAOMI", result["message"])
        self.assertIn("Дахуа", result["message"])
        self.assertIn('style="color:#dc2626;font-weight:800"', result["message_html"])

    def test_asian_brands_passes_when_not_found(self):
        tmp, base = make_result_dir(make_ocr_log({"frame_001.jpg": ["Основной бренд"]}))
        try:
            result = evaluate_asian_brands(base)
        finally:
            tmp.cleanup()

        self.assertEqual(result["status"], "pass")
        self.assertEqual(result["message"], "Азиатских брендов не обнаружено")

    def test_asian_brand_cv_detection_warns_without_ocr(self):
        tmp, base = make_result_dir(make_ocr_log({"frame_001.jpg": ["нейтральный текст"]}))
        fake_cv = {
            "cv_enabled": True,
            "cv_model_path": "model.pt",
            "cv_error": "",
            "cv_brand_mentions": [
                {
                    "brand": "XIAOMI",
                    "seconds": ["1сек."],
                    "frames": ["frame_001.jpg"],
                    "sources": ["cv"],
                    "cv_labels": ["XIAOMI"],
                    "duration_sec": 1,
                }
            ],
            "cv_brand_bearing_objects": [],
        }
        try:
            with patch("app.secondary_advertisers.analyze_secondary_advertisers_from_cv", return_value=fake_cv):
                result = evaluate_asian_brands(base)
        finally:
            tmp.cleanup()

        self.assertEqual(result["status"], "warning")
        self.assertIn("XIAOMI", result["message"])

    def test_evaluates_item_id_272_inside_view_model(self):
        tmp, base = make_result_dir(make_ocr_log({"frame_001.jpg": ["70mai"]}))
        view_model = {
            "ok": True,
            "blocks": [{"name": "Видеоряд", "items": [{"id": "272", "number": "272", "text": "asia"}]}],
        }
        try:
            evaluated = evaluate_approval_view_model(view_model, base)
        finally:
            tmp.cleanup()

        item = evaluated["blocks"][0]["items"][0]
        self.assertEqual(item["status"], "warning")
        self.assertIn("70mai", item["message"])

    def test_cv_generic_labels_include_packaging_and_labels(self):
        normalized = {label.lower() for label in BRAND_BEARING_CV_LABELS}
        self.assertIn("product packaging", normalized)
        self.assertIn("label", normalized)
        self.assertIn("brand logo", normalized)


if __name__ == "__main__":
    unittest.main()
